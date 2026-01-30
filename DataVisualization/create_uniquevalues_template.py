"""
UniqueValues Template Generator
Creates the UniqueValuesColorSettings.xlsx file structure
This file will be populated by Script 02 with color mapping tables

Author: RhinoGuire
Version: 1.0
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Install with: pip install openpyxl")
    exit()

import os

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "ColorMaps"

# Header styling
title_font = Font(bold=True, size=14, color="FFFFFF")
title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Create title and instructions
ws['A1'] = 'Unique Values and Color Settings'
ws['A1'].font = Font(bold=True, size=16)

ws['A3'] = 'Instructions:'
ws['A3'].font = Font(bold=True, size=12)
ws['A4'] = '1. Run Script 02 to scan your Rhino objects and auto-generate color mapping tables below'
ws['A5'] = '2. Each table corresponds to one metadata key from your objects'
ws['A6'] = '3. Fill in the R, G, B, A values for each unique value (A=Alpha is optional, defaults to 255)'
ws['A7'] = '4. Run Script 03 to visualize objects using these color assignments'
ws['A8'] = ''
ws['A9'] = 'Color Format: R,G,B,A where each value is 0-255'
ws['A10'] = 'Example: 190,190,190,255 = opaque gray | 255,0,0,128 = semi-transparent red'

ws['A12'] = 'Note: Tables will be arranged in a 6-column grid below (scroll right to see more tables)'
ws['A12'].font = Font(italic=True, color="666666")

# Style instructions
for row in range(3, 13):
    ws[f'A{row}'].alignment = Alignment(wrap_text=True)

# Set column width for instructions
ws.column_dimensions['A'].width = 80

# Mark where tables will start
ws['A15'] = '← Color mapping tables will appear below this line when you run Script 02'
ws['A15'].font = Font(italic=True, color="999999")
ws['A15'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

# Merge cells for better visibility
ws.merge_cells('A15:F15')

# Save the file
script_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(script_dir, 'UniqueValuesColorSettings.xlsx')
wb.save(output_path)

print(f"UniqueValues template created successfully: {output_path}")
print("\nThis template will be populated by Script 02 with:")
print("- Color mapping tables in 6-column grid layout")
print("- One table per unique metadata key found in objects")
print("- Columns: Value | R | G | B | A")
