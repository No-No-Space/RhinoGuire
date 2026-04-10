#! python3
# # -*- coding: utf-8 -*-
# __title__ = "Chivito"                           # Name of the button displayed in Revit UI
# __doc__ = """Version = 0.5
# Date    = 2026-02-14
# Author: Aquelon - aquelon@pm.me 
# _____________________________________________________________________
# Description:
# Data Visualization Tool - Unified Interface (Python 3 / Excel Version)
# A single GUI for the complete metadata visualization workflow using Excel files.
# FULLY SELF-CONTAINED - No external script dependencies
# _____________________________________________________________________
# How-to:
# -> Run the script in Rhino 8 (RunPythonScript), a modeless window opens
# -> Step 1: Click "Select Excel & Apply Keys", pick an Excel template, select objects
# -> Step 2: Click "Scan & Generate Excel", select objects, save the color mapping file
# -> Open the exported Excel in a spreadsheet editor and fill in the R, G, B, A columns
# -> Step 3: Click "Open Color Manager", select objects, pick the color Excel file
# -> In the Color Manager: select a key from the dropdown, click "Update Colors"
# -> Use "Export Legend as PNG" or "Capture Viewport as PNG" for presentations
# -> The window is non-blocking, you can interact with Rhino while it is open
# _____________________________________________________________________
# Last update:
# - [14.02.2026] - 0.5 RELEASE
# _____________________________________________________________________
# To-Do:
# - UI needs improvement, it is functional at the moment, but default look.
# - Default folder locations need to be updated to neutral locations (desktop or documents) instead of script folder.
# - Add a button to create an Excel template, before it was a separate script, but it may be more intuitive to have it integrated in the workflow, this way the user can create the template and fill it with colors before applying it to the objects.
# The script window needs to be closed before editing the Excel file exported. Correct this, so the file is freed after export.
# - Correct that the ViewCaptureToFile allows for interaction with the menu from Rhino, so the user cna define the settings of the capture.
# _____________________________________________________________________
# r: openpyxl

import rhinoscriptsyntax as rs
import scriptcontext as sc
import Rhino
import os
from collections import defaultdict

# Import openpyxl for Excel operations
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font as ExcelFont
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available. Please install it using: pip install openpyxl")

# Import Eto forms
import Eto.Drawing as drawing
import Eto.Forms as forms

# Import System.Drawing for PNG export
import System.Drawing as sdrawing


# ============================================================================
# SCRIPT 01 FUNCTIONS - Read keys from Excel and apply to objects
# ============================================================================

def read_keys_from_excel(excel_path):
    """Read key definitions from Excel file

    Expected format:
    Column A: Key Name
    Column B: Default Value
    """
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        keys = {}
        # Skip header row, start from row 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If key name exists
                key_name = str(row[0]).strip()
                default_value = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if key_name:
                    keys[key_name] = default_value

        wb.close()
        return keys
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return None


def apply_keys_to_objects(objects, keys):
    """Apply keys to objects without overwriting"""
    stats = {'objects_processed': 0, 'keys_added': 0, 'keys_skipped': 0}
    for obj in objects:
        stats['objects_processed'] += 1
        for key_name, default_value in keys.items():
            if rs.GetUserText(obj, key_name) is None:
                rs.SetUserText(obj, key_name, default_value)
                stats['keys_added'] += 1
            else:
                stats['keys_skipped'] += 1
    return stats


# ============================================================================
# SCRIPT 02 FUNCTIONS - Collect unique values and generate Excel
# ============================================================================

def collect_unique_values(objects):
    """Scan objects for unique key-value pairs"""
    key_values = defaultdict(set)
    for obj in objects:
        keys = rs.GetUserText(obj)
        if keys:
            for key in keys:
                value = rs.GetUserText(obj, key)
                if value is not None and value != '':
                    key_values[key].add(value)
    return {key: sorted(list(values)) for key, values in key_values.items()}


def write_color_map_excel(excel_path, key_value_dict):
    """Write color mapping Excel file, preserving existing color values.

    If the target file already exists, existing RGB/A values are preserved
    for key-value pairs that still exist. Only entries no longer present
    in the scan are removed.

    Format:
    Columns: Key | Value | R | G | B | A
    """
    # Read existing color values if the file already exists
    existing_colors = {}
    if os.path.exists(excel_path):
        try:
            old_wb = load_workbook(excel_path)
            old_ws = old_wb.active
            for row in old_ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 6 and row[0] and row[1]:
                    old_key = str(row[0]).strip()
                    old_value = str(row[1]).strip()
                    old_r = row[2] if row[2] is not None else ''
                    old_g = row[3] if row[3] is not None else ''
                    old_b = row[4] if row[4] is not None else ''
                    old_a = row[5] if row[5] is not None else ''
                    existing_colors[(old_key, old_value)] = (old_r, old_g, old_b, old_a)
            old_wb.close()
            print(f"Found {len(existing_colors)} existing color definitions to preserve.")
        except Exception as e:
            print(f"Could not read existing file, creating new: {e}")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Color Mapping"

        # Write headers
        headers = ['Key', 'Value', 'R', 'G', 'B', 'A']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = ExcelFont(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data rows, preserving existing colors
        row_idx = 2
        preserved = 0
        for key_name in sorted(key_value_dict.keys()):
            for value in key_value_dict[key_name]:
                ws.cell(row=row_idx, column=1, value=key_name)
                ws.cell(row=row_idx, column=2, value=value)

                # Check if color values exist for this key-value pair
                old_rgba = existing_colors.get((key_name, value))
                if old_rgba:
                    ws.cell(row=row_idx, column=3, value=old_rgba[0])
                    ws.cell(row=row_idx, column=4, value=old_rgba[1])
                    ws.cell(row=row_idx, column=5, value=old_rgba[2])
                    ws.cell(row=row_idx, column=6, value=old_rgba[3])
                    preserved += 1
                else:
                    ws.cell(row=row_idx, column=3, value='')
                    ws.cell(row=row_idx, column=4, value='')
                    ws.cell(row=row_idx, column=5, value='')
                    ws.cell(row=row_idx, column=6, value='')

                row_idx += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(excel_path)
        wb.close()

        total_rows = row_idx - 2
        if preserved > 0:
            print(f"Preserved {preserved} existing color definitions out of {total_rows} entries.")
        return total_rows
    except Exception as e:
        print(f"Error writing Excel: {e}")
        return 0


# ============================================================================
# SCRIPT 03 FUNCTIONS - Apply colors from Excel mapping
# ============================================================================

def read_color_map_from_excel(excel_path):
    """Read color mappings from Excel file

    Expected format:
    Columns: Key | Value | R | G | B | A
    """
    color_map = {}
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        # Read data rows (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 6:
                key = str(row[0]).strip() if row[0] else ''
                value = str(row[1]).strip() if row[1] else ''
                r = row[2]
                g = row[3]
                b = row[4]
                a = row[5]

                if key and value:
                    try:
                        r_val = max(0, min(255, int(r) if r else 0))
                        g_val = max(0, min(255, int(g) if g else 0))
                        b_val = max(0, min(255, int(b) if b else 0))
                        a_val = max(0, min(255, int(a) if a else 255))

                        if key not in color_map:
                            color_map[key] = {}
                        color_map[key][value] = (r_val, g_val, b_val, a_val)
                    except ValueError:
                        continue

        wb.close()
        return color_map
    except Exception as e:
        print(f"Error reading color Excel: {e}")
        return None


def analyze_objects_for_key(objects, active_key):
    """Analyze objects for unique values"""
    value_objects = {}
    objects_without_key = []
    for obj in objects:
        value = rs.GetUserText(obj, active_key)
        if value is None or value == '':
            objects_without_key.append(obj)
        else:
            if value not in value_objects:
                value_objects[value] = []
            value_objects[value].append(obj)
    return {'value_objects': value_objects, 'objects_without_key': objects_without_key}


def apply_colors_to_objects(objects, active_key, color_map, default_color):
    """Apply colors based on metadata - optimized with RhinoCommon direct access.

    Key optimizations over rhinoscriptsyntax:
    - Single pass: reads user text and applies color in one loop
    - Redraw suppressed during batch operation
    - Pre-computed System.Drawing.Color objects
    - Direct attribute modification (bypasses rs.ObjectColor overhead)
    """
    stats = {
        'colored': 0,
        'default': 0,
        'missing_color_definition': [],
        'unused_color_definitions': []
    }

    key_color_map = color_map.get(active_key, {})

    # Pre-compute System.Drawing.Color for each value
    color_cache = {}
    for value, rgba in key_color_map.items():
        color_cache[value] = sdrawing.Color.FromArgb(rgba[0], rgba[1], rgba[2])
    default_sys_color = sdrawing.Color.FromArgb(
        default_color[0], default_color[1], default_color[2])
    color_from_obj = Rhino.DocObjects.ObjectColorSource.ColorFromObject

    used_values = set()
    missing_by_value = {}

    # Disable viewport redraws during batch operation
    sc.doc.Views.RedrawEnabled = False
    try:
        for obj_id in objects:
            rhino_obj = sc.doc.Objects.FindId(rs.coerceguid(obj_id))
            if not rhino_obj:
                continue

            value = rhino_obj.Attributes.GetUserString(active_key)
            attrs = rhino_obj.Attributes.Duplicate()
            attrs.ColorSource = color_from_obj

            if not value:
                attrs.ObjectColor = default_sys_color
                sc.doc.Objects.ModifyAttributes(rhino_obj.Id, attrs, True)
                stats['default'] += 1
            else:
                used_values.add(value)
                sys_color = color_cache.get(value)
                if sys_color:
                    attrs.ObjectColor = sys_color
                    sc.doc.Objects.ModifyAttributes(rhino_obj.Id, attrs, True)
                    stats['colored'] += 1
                else:
                    attrs.ObjectColor = default_sys_color
                    sc.doc.Objects.ModifyAttributes(rhino_obj.Id, attrs, True)
                    if value not in missing_by_value:
                        missing_by_value[value] = []
                    missing_by_value[value].append(obj_id)
    finally:
        sc.doc.Views.RedrawEnabled = True

    stats['missing_color_definition'] = [
        (v, objs) for v, objs in missing_by_value.items()]

    excel_values = set(key_color_map.keys())
    unused = excel_values - used_values
    stats['unused_color_definitions'] = sorted(list(unused))

    return stats


# ============================================================================
# COLOR MANAGER DIALOG (SCRIPT 03)
# ============================================================================

class ColorManagerDialog(forms.Form):
    """Color Manager GUI (modeless - allows Rhino interaction)"""

    def __init__(self, excel_path, objects):
        super(ColorManagerDialog, self).__init__()

        self.excel_path = excel_path
        self.objects = objects
        self.color_map = None
        self.active_key = None
        self.default_color = (128, 128, 128)
        self.problem_objects = []

        self.load_color_map()

        self.Title = "Color Manager"
        self.Padding = drawing.Padding(10)
        self.Resizable   = True
        self.MinimumSize = drawing.Size(480, 540)
        self.ClientSize  = drawing.Size(520, 640)
        self.Topmost = True  # Keep on top

        self.create_controls()
        self.create_layout()

    def load_color_map(self):
        self.color_map = read_color_map_from_excel(self.excel_path)
        if not self.color_map:
            self.color_map = {}
            forms.MessageBox.Show("Error loading Excel file", "Error")

    def create_controls(self):
        self.key_label = forms.Label()
        self.key_label.Text = "Select Metadata Key:"

        self.key_dropdown = forms.DropDown()
        available_keys = sorted(self.color_map.keys())
        for key in available_keys:
            self.key_dropdown.Items.Add(key)
        if available_keys:
            self.key_dropdown.SelectedIndex = 0
            self.active_key = available_keys[0]
        self.key_dropdown.SelectedIndexChanged += self.on_key_changed

        self.default_color_label = forms.Label()
        self.default_color_label.Text = "Default Color:"

        self.default_color_picker = forms.ColorPicker()
        self.default_color_picker.Value = drawing.Color.FromArgb(128, 128, 128)
        self.default_color_picker.ValueChanged += self.on_default_color_changed

        self.legend_label = forms.Label()
        self.legend_label.Text = "Color Legend:"

        self.legend_panel = forms.Panel()
        self.legend_panel.Size = drawing.Size(450, 200)
        self.legend_panel.BackgroundColor = drawing.Colors.White
        self.update_legend()

        self.warnings_label = forms.Label()
        self.warnings_label.Text = "Warnings & Errors:"

        self.warnings_text = forms.TextArea()
        self.warnings_text.ReadOnly = True
        self.warnings_text.Size = drawing.Size(450, 100)
        self.warnings_text.Text = "Click 'Update Colors' to analyze..."

        self.update_button = forms.Button()
        self.update_button.Text = "Update Colors"
        self.update_button.Click += self.on_update_clicked

        self.select_button = forms.Button()
        self.select_button.Text = "Select Problem Objects"
        self.select_button.Click += self.on_select_clicked
        self.select_button.Enabled = False

        self.export_legend_button = forms.Button()
        self.export_legend_button.Text = "Export Legend as PNG"
        self.export_legend_button.Click += self.on_export_legend_clicked

        self.capture_viewport_button = forms.Button()
        self.capture_viewport_button.Text = "Capture Viewport as PNG"
        self.capture_viewport_button.Click += self.on_capture_viewport_clicked

        self.close_button = forms.Button()
        self.close_button.Text = "Close"
        self.close_button.Click += self.on_close_clicked

        self.status_label = forms.Label()
        self.status_label.Text = f"Ready - {len(self.objects)} objects"

    def create_layout(self):
        layout = forms.DynamicLayout()
        layout.Spacing = drawing.Size(5, 5)
        layout.AddRow(self.key_label)
        layout.AddRow(self.key_dropdown)
        layout.AddRow(None)
        layout.AddRow(self.default_color_label)
        layout.AddRow(self.default_color_picker)
        layout.AddRow(None)
        layout.AddRow(self.legend_label)
        layout.AddRow(self.legend_panel)
        layout.AddRow(None)
        layout.AddRow(self.warnings_label)
        layout.AddRow(self.warnings_text)
        layout.AddRow(None)

        button_layout = forms.DynamicLayout()
        button_layout.Spacing = drawing.Size(5, 5)
        button_layout.AddRow(self.update_button, self.select_button)
        button_layout.AddRow(self.export_legend_button, self.capture_viewport_button, self.close_button)
        layout.AddRow(button_layout)
        layout.AddRow(self.status_label)

        self.Content = layout

    def update_legend(self):
        if not self.active_key:
            return
        legend_layout = forms.DynamicLayout()
        legend_layout.Spacing = drawing.Size(5, 5)
        legend_layout.Padding = drawing.Padding(10)

        title = forms.Label()
        title.Text = self.active_key
        title.Font = drawing.Font(drawing.SystemFont.Bold, 12)
        legend_layout.AddRow(title)
        legend_layout.AddRow(None)

        key_colors = self.color_map.get(self.active_key, {})
        for value in sorted(key_colors.keys()):
            color_rgb = key_colors[value]

            color_panel = forms.Panel()
            color_panel.Size = drawing.Size(30, 20)
            color_panel.BackgroundColor = drawing.Color.FromArgb(color_rgb[0], color_rgb[1], color_rgb[2])

            value_label = forms.Label()
            value_label.Text = value

            rgb_text = f"RGB: {color_rgb[0]},{color_rgb[1]},{color_rgb[2]}"
            rgb_label = forms.Label()
            rgb_label.Text = rgb_text
            rgb_label.Font = drawing.Font(drawing.SystemFont.Default, 9)
            rgb_label.TextColor = drawing.Colors.Gray

            row_layout = forms.DynamicLayout()
            row_layout.Spacing = drawing.Size(5, 5)
            row_layout.AddRow(color_panel, value_label, rgb_label)
            legend_layout.AddRow(row_layout)

        self.legend_panel.Content = legend_layout

    def update_warnings(self, stats):
        warnings = []
        self.problem_objects = []

        if stats['missing_color_definition']:
            warnings.append("ERROR: Values in objects but missing from Excel:")
            warnings.append("=" * 50)
            for value, obj_list in stats['missing_color_definition']:
                warnings.append(f"  '{value}' ({len(obj_list)} objects)")
                self.problem_objects.extend(obj_list)
            warnings.append("")

        if stats['unused_color_definitions']:
            warnings.append("WARNING: Values in Excel but not in selection:")
            warnings.append("=" * 50)
            for value in stats['unused_color_definitions']:
                warnings.append(f"  '{value}'")
            warnings.append("")

        if stats['default'] > 0:
            warnings.append(f"INFO: {stats['default']} objects without '{self.active_key}' key")
            warnings.append("")

        if not warnings:
            warnings.append("SUCCESS: All objects colored correctly!")
            warnings.append(f"  {stats['colored']} objects colored")

        self.warnings_text.Text = "\n".join(warnings)
        self.select_button.Enabled = len(self.problem_objects) > 0

    def on_key_changed(self, sender, e):
        selected = self.key_dropdown.SelectedValue
        if selected:
            self.active_key = str(selected)
            self.update_legend()
            self.warnings_text.Text = "Click 'Update Colors' to analyze..."
            self.select_button.Enabled = False

    def on_default_color_changed(self, sender, e):
        color = self.default_color_picker.Value
        self.default_color = (color.R, color.G, color.B)

    def on_update_clicked(self, sender, e):
        if not self.active_key:
            return

        self.status_label.Text = "Applying colors..."
        self.load_color_map()
        self.update_legend()

        stats = apply_colors_to_objects(self.objects, self.active_key, self.color_map, self.default_color)
        sc.doc.Views.Redraw()

        self.update_warnings(stats)
        self.status_label.Text = f"Updated! Colored: {stats['colored']}, Default: {stats['default']}"

    def on_select_clicked(self, sender, e):
        if self.problem_objects:
            rs.UnselectAllObjects()
            rs.SelectObjects(self.problem_objects)
            forms.MessageBox.Show(f"Selected {len(self.problem_objects)} objects with missing colors", "Objects Selected")

    def on_export_legend_clicked(self, sender, e):
        """Export the current legend as a PNG image"""
        if not self.active_key or not self.color_map:
            forms.MessageBox.Show("No legend to export. Select a key first.", "Export Legend")
            return

        key_colors = self.color_map.get(self.active_key, {})
        if not key_colors:
            forms.MessageBox.Show(f"No color definitions for key '{self.active_key}'.", "Export Legend")
            return

        # Ask for save location
        png_path = rs.SaveFileName("Save Legend as PNG", "PNG Files (*.png)|*.png||",
                                   filename=f"Legend_{self.active_key}.png")
        if not png_path:
            return

        try:
            # Layout constants
            padding = 20
            swatch_size = 24
            swatch_gap = 10
            row_height = 30
            title_height = 35
            sorted_values = sorted(key_colors.keys())

            # Measure text to determine image width
            measure_bmp = sdrawing.Bitmap(1, 1)
            measure_gfx = sdrawing.Graphics.FromImage(measure_bmp)
            title_font = sdrawing.Font("Arial", 14, sdrawing.FontStyle.Bold)
            value_font = sdrawing.Font("Arial", 10)
            rgb_font = sdrawing.Font("Arial", 9)

            max_text_width = 0
            for value in sorted_values:
                color_rgb = key_colors[value]
                value_text = value
                rgb_text = f"  RGB({color_rgb[0]}, {color_rgb[1]}, {color_rgb[2]})"
                full_text = value_text + rgb_text
                text_size = measure_gfx.MeasureString(full_text, value_font)
                if text_size.Width > max_text_width:
                    max_text_width = text_size.Width

            measure_gfx.Dispose()
            measure_bmp.Dispose()

            # Calculate image dimensions
            img_width = int(padding + swatch_size + swatch_gap + max_text_width + padding + 20)
            img_width = max(img_width, 300)
            img_height = padding + title_height + len(sorted_values) * row_height + padding

            # Create bitmap and draw
            bmp = sdrawing.Bitmap(img_width, img_height)
            gfx = sdrawing.Graphics.FromImage(bmp)
            gfx.SmoothingMode = sdrawing.Drawing2D.SmoothingMode.AntiAlias
            gfx.TextRenderingHint = sdrawing.Text.TextRenderingHint.AntiAlias

            # White background
            gfx.Clear(sdrawing.Color.White)

            # Draw title
            gfx.DrawString(self.active_key, title_font,
                           sdrawing.Brushes.Black,
                           sdrawing.PointF(padding, padding))

            # Draw each legend entry
            y = padding + title_height
            for value in sorted_values:
                color_rgb = key_colors[value]

                # Color swatch
                brush = sdrawing.SolidBrush(
                    sdrawing.Color.FromArgb(color_rgb[0], color_rgb[1], color_rgb[2])
                )
                gfx.FillRectangle(brush,
                                  padding, int(y), swatch_size, swatch_size)
                # Swatch border
                gfx.DrawRectangle(sdrawing.Pens.Gray,
                                  padding, int(y), swatch_size, swatch_size)
                brush.Dispose()

                # Value text
                text_x = padding + swatch_size + swatch_gap
                gfx.DrawString(value, value_font,
                               sdrawing.Brushes.Black,
                               sdrawing.PointF(text_x, y))

                # RGB text
                rgb_text = f"  RGB({color_rgb[0]}, {color_rgb[1]}, {color_rgb[2]})"
                value_size = gfx.MeasureString(value, value_font)
                gfx.DrawString(rgb_text, rgb_font,
                               sdrawing.Brushes.Gray,
                               sdrawing.PointF(text_x + value_size.Width, y + 2))

                y += row_height

            # Save
            bmp.Save(png_path, sdrawing.Imaging.ImageFormat.Png)

            # Cleanup
            title_font.Dispose()
            value_font.Dispose()
            rgb_font.Dispose()
            gfx.Dispose()
            bmp.Dispose()

            self.status_label.Text = "Legend exported as PNG"
            forms.MessageBox.Show(f"Legend exported!\n\n{os.path.basename(png_path)}", "Export Complete")

        except Exception as ex:
            print(f"Error exporting legend: {ex}")
            forms.MessageBox.Show(f"Error exporting legend:\n{ex}", "Export Error")

    def on_capture_viewport_clicked(self, sender, e):
        """Open Rhino's ViewCaptureToFile dialog for viewport capture."""
        self.status_label.Text = "Opening viewport capture settings..."
        Rhino.RhinoApp.RunScript("ViewCaptureToFile", False)
        self.status_label.Text = "Viewport capture complete"

    def on_close_clicked(self, sender, e):
        self.Close()


# ============================================================================
# MAIN UNIFIED INTERFACE
# ============================================================================

class DataVisualizationTool(forms.Form):
    """Main unified interface - modeless, allows Rhino interaction"""

    def __init__(self):
        super(DataVisualizationTool, self).__init__()

        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_folder = os.path.join(self.script_dir, "_ExcelOutput")

        # Create output folder if it doesn't exist
        if not os.path.exists(self.excel_folder):
            try:
                os.makedirs(self.excel_folder)
            except:
                pass

        self.Title = "Data Visualization Tool (Excel)"
        self.Padding = drawing.Padding(15)
        self.Resizable   = True
        self.MinimumSize = drawing.Size(420, 400)
        self.ClientSize  = drawing.Size(480, 480)
        self.Topmost = True  # KEEP ON TOP

        self.create_controls()
        self.create_layout()

    def create_controls(self):
        self.header = forms.Label()
        self.header.Text = "Rhino Metadata Visualization"
        self.header.Font = drawing.Font(drawing.SystemFont.Bold, 12)

        self.step1_label = forms.Label()
        self.step1_label.Text = "Step 1: Initialize Keys"
        self.step1_label.Font = drawing.Font(drawing.SystemFont.Bold, 10)

        self.step1_desc = forms.Label()
        self.step1_desc.Text = "Apply metadata keys from Excel to objects"
        self.step1_desc.TextColor = drawing.Colors.Gray

        self.step1_button = forms.Button()
        self.step1_button.Text = "Select Excel & Apply Keys"
        self.step1_button.Click += self.on_step1

        self.step2_label = forms.Label()
        self.step2_label.Text = "Step 2: Extract Unique Values"
        self.step2_label.Font = drawing.Font(drawing.SystemFont.Bold, 10)

        self.step2_desc = forms.Label()
        self.step2_desc.Text = "Scan objects and generate color Excel"
        self.step2_desc.TextColor = drawing.Colors.Gray

        self.step2_button = forms.Button()
        self.step2_button.Text = "Scan & Generate Excel"
        self.step2_button.Click += self.on_step2

        self.step3_label = forms.Label()
        self.step3_label.Text = "Step 3: Visualize with Colors"
        self.step3_label.Font = drawing.Font(drawing.SystemFont.Bold, 10)

        self.step3_desc = forms.Label()
        self.step3_desc.Text = "Interactive color visualization"
        self.step3_desc.TextColor = drawing.Colors.Gray

        self.step3_button = forms.Button()
        self.step3_button.Text = "Open Color Manager"
        self.step3_button.Click += self.on_step3

        self.status = forms.Label()
        self.status.Text = "Ready"
        self.status.TextColor = drawing.Colors.Blue

        self.close_button = forms.Button()
        self.close_button.Text = "Close"
        self.close_button.Click += self.on_close

    def create_layout(self):
        layout = forms.DynamicLayout()
        layout.Spacing = drawing.Size(8, 8)
        layout.AddRow(self.header)
        layout.AddRow(None)
        layout.AddRow(self.step1_label)
        layout.AddRow(self.step1_desc)
        layout.AddRow(self.step1_button)
        layout.AddRow(None)
        layout.AddRow(self.step2_label)
        layout.AddRow(self.step2_desc)
        layout.AddRow(self.step2_button)
        layout.AddRow(None)
        layout.AddRow(self.step3_label)
        layout.AddRow(self.step3_desc)
        layout.AddRow(self.step3_button)
        layout.AddRow(None)
        layout.AddRow(self.status)
        layout.AddRow(self.close_button)
        self.Content = layout

    def on_step1(self, sender, e):
        excel_path = rs.OpenFileName("Select Excel File", "Excel Files (*.xlsx)|*.xlsx||",
                                     folder=self.excel_folder if os.path.exists(self.excel_folder) else None)
        if not excel_path:
            return
        objects = rs.GetObjects("Select objects", preselect=True)
        if not objects:
            return
        keys = read_keys_from_excel(excel_path)
        if not keys:
            forms.MessageBox.Show("No keys found in Excel file", "Error")
            return
        stats = apply_keys_to_objects(objects, keys)
        self.status.Text = f"Step 1: {stats['keys_added']} keys added"
        forms.MessageBox.Show(f"Keys applied!\n\nObjects: {stats['objects_processed']}\nKeys added: {stats['keys_added']}\nSkipped: {stats['keys_skipped']}", "Complete")
        self.BringToFront()  # Return focus

    def on_step2(self, sender, e):
        objects = rs.GetObjects("Select objects", preselect=True)
        if not objects:
            return
        kvd = collect_unique_values(objects)
        if not kvd:
            forms.MessageBox.Show("No metadata found", "Error")
            return
        excel_path = rs.SaveFileName("Save Excel File", "Excel Files (*.xlsx)|*.xlsx||",
                                     folder=self.excel_folder if os.path.exists(self.excel_folder) else None,
                                     filename="UniqueValuesColorSettings.xlsx")
        if not excel_path:
            return
        file_existed = os.path.exists(excel_path)
        rows = write_color_map_excel(excel_path, kvd)
        self.status.Text = f"Step 2: {rows} values exported"
        msg = f"Excel generated!\n\nKeys: {len(kvd)}\nValues: {rows}"
        if file_existed:
            msg += "\n\nExisting color definitions were preserved."
        msg += "\n\nNext: Fill RGB colors in Excel"
        forms.MessageBox.Show(msg, "Complete")
        self.BringToFront()  # Return focus

    def on_step3(self, sender, e):
        objects = rs.GetObjects("Select objects", preselect=True)
        if not objects:
            return
        excel_path = rs.OpenFileName("Select color Excel file", "Excel Files (*.xlsx)|*.xlsx||",
                                     folder=self.excel_folder if os.path.exists(self.excel_folder) else None)
        if not excel_path:
            return
        # Launch Color Manager as modeless form
        color_dlg = ColorManagerDialog(excel_path, objects)
        color_dlg.Owner = Rhino.UI.RhinoEtoApp.MainWindow
        color_dlg.Show()
        self.status.Text = "Step 3: Color Manager opened"

    def on_close(self, sender, e):
        self.Close()


def main():
    """Main entry point"""
    if not EXCEL_AVAILABLE:
        print("="*60)
        print("ERROR: openpyxl library is not installed")
        print("="*60)
        print("\nThis script requires the openpyxl library to work with Excel files.")
        print("Please install it using the following command:\n")
        print("  pip install openpyxl\n")
        print("If you're using Rhino 8's Python 3, you may need to:")
        print("  1. Open Windows Command Prompt (cmd)")
        print("  2. Navigate to Rhino's Python directory")
        print("  3. Run: python -m pip install openpyxl")
        print("="*60)
        return

    dialog = DataVisualizationTool()
    dialog.Owner = Rhino.UI.RhinoEtoApp.MainWindow
    dialog.Show()


if __name__ == "__main__":
    main()
