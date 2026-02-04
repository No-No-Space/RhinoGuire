#! python3

"""
Data Visualization Tool - Unified Interface (Python 3 / Excel Version)
A single GUI for the complete metadata visualization workflow using Excel files.

FULLY SELF-CONTAINED - No external script dependencies

Author: RhinoGuire
Version: 2.0 (Python 3 + Excel)
"""

# r: openpyxl

import rhinoscriptsyntax as rs
import scriptcontext as sc
import os
from collections import defaultdict

# Import openpyxl for Excel operations
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font as ExcelFont
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available. Please install it using: pip install openpyxl")

# Import Eto forms
import Eto.Drawing as drawing
import Eto.Forms as forms


# ============================================================================
# SCRIPT 01 FUNCTIONS - Read keys from Excel and apply to objects
# ============================================================================

def read_keys_from_excel(excel_path):
    """Read key definitions from Excel file

    Expected format:
    Column A: Key Name
    Column B: Default Value
    """
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        keys = {}
        # Skip header row, start from row 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If key name exists
                key_name = str(row[0]).strip()
                default_value = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if key_name:
                    keys[key_name] = default_value

        wb.close()
        return keys
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return None


def apply_keys_to_objects(objects, keys):
    """Apply keys to objects without overwriting"""
    stats = {'objects_processed': 0, 'keys_added': 0, 'keys_skipped': 0}
    for obj in objects:
        stats['objects_processed'] += 1
        for key_name, default_value in keys.items():
            if rs.GetUserText(obj, key_name) is None:
                rs.SetUserText(obj, key_name, default_value)
                stats['keys_added'] += 1
            else:
                stats['keys_skipped'] += 1
    return stats


# ============================================================================
# SCRIPT 02 FUNCTIONS - Collect unique values and generate Excel
# ============================================================================

def collect_unique_values(objects):
    """Scan objects for unique key-value pairs"""
    key_values = defaultdict(set)
    for obj in objects:
        keys = rs.GetUserText(obj)
        if keys:
            for key in keys:
                value = rs.GetUserText(obj, key)
                if value is not None and value != '':
                    key_values[key].add(value)
    return {key: sorted(list(values)) for key, values in key_values.items()}


def write_color_map_excel(excel_path, key_value_dict):
    """Write color mapping Excel file

    Format:
    Columns: Key | Value | R | G | B | A
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Color Mapping"

        # Write headers
        headers = ['Key', 'Value', 'R', 'G', 'B', 'A']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = ExcelFont(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data rows
        row_idx = 2
        for key_name in sorted(key_value_dict.keys()):
            for value in key_value_dict[key_name]:
                ws.cell(row=row_idx, column=1, value=key_name)
                ws.cell(row=row_idx, column=2, value=value)
                ws.cell(row=row_idx, column=3, value='')  # R
                ws.cell(row=row_idx, column=4, value='')  # G
                ws.cell(row=row_idx, column=5, value='')  # B
                ws.cell(row=row_idx, column=6, value='')  # A
                row_idx += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(excel_path)
        wb.close()
        return row_idx - 2  # Return number of data rows
    except Exception as e:
        print(f"Error writing Excel: {e}")
        return 0


# ============================================================================
# SCRIPT 03 FUNCTIONS - Apply colors from Excel mapping
# ============================================================================

def read_color_map_from_excel(excel_path):
    """Read color mappings from Excel file

    Expected format:
    Columns: Key | Value | R | G | B | A
    """
    color_map = {}
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        # Read data rows (skip header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 6:
                key = str(row[0]).strip() if row[0] else ''
                value = str(row[1]).strip() if row[1] else ''
                r = row[2]
                g = row[3]
                b = row[4]
                a = row[5]

                if key and value:
                    try:
                        r_val = max(0, min(255, int(r) if r else 0))
                        g_val = max(0, min(255, int(g) if g else 0))
                        b_val = max(0, min(255, int(b) if b else 0))
                        a_val = max(0, min(255, int(a) if a else 255))

                        if key not in color_map:
                            color_map[key] = {}
                        color_map[key][value] = (r_val, g_val, b_val, a_val)
                    except ValueError:
                        continue

        wb.close()
        return color_map
    except Exception as e:
        print(f"Error reading color Excel: {e}")
        return None


def analyze_objects_for_key(objects, active_key):
    """Analyze objects for unique values"""
    value_objects = {}
    objects_without_key = []
    for obj in objects:
        value = rs.GetUserText(obj, active_key)
        if value is None or value == '':
            objects_without_key.append(obj)
        else:
            if value not in value_objects:
                value_objects[value] = []
            value_objects[value].append(obj)
    return {'value_objects': value_objects, 'objects_without_key': objects_without_key}


def apply_colors_to_objects(objects, active_key, color_map, default_color):
    """Apply colors based on metadata"""
    stats = {
        'colored': 0,
        'default': 0,
        'missing_color_definition': [],
        'unused_color_definitions': []
    }

    key_color_map = color_map.get(active_key, {})
    analysis = analyze_objects_for_key(objects, active_key)
    value_objects = analysis['value_objects']
    objects_without_key = analysis['objects_without_key']
    used_values = set()

    for value, obj_list in value_objects.items():
        used_values.add(value)
        color = key_color_map.get(value)
        if color:
            for obj in obj_list:
                rs.ObjectColor(obj, (color[0], color[1], color[2]))
                stats['colored'] += 1
        else:
            for obj in obj_list:
                rs.ObjectColor(obj, default_color)
            stats['missing_color_definition'].append((value, obj_list))

    for obj in objects_without_key:
        rs.ObjectColor(obj, default_color)
        stats['default'] += 1

    excel_values = set(key_color_map.keys())
    unused = excel_values - used_values
    stats['unused_color_definitions'] = sorted(list(unused))

    return stats


# ============================================================================
# COLOR MANAGER DIALOG (SCRIPT 03)
# ============================================================================

class ColorManagerDialog(forms.Dialog[bool]):
    """Color Manager GUI"""

    def __init__(self, excel_path, objects):
        # Initialize base Dialog class first
        super(ColorManagerDialog, self).__init__()

        self.excel_path = excel_path
        self.objects = objects
        self.color_map = None
        self.active_key = None
        self.default_color = (128, 128, 128)
        self.problem_objects = []

        self.load_color_map()

        self.Title = "Color Manager"
        self.Padding = drawing.Padding(10)
        self.Resizable = True
        self.Size = drawing.Size(500, 600)
        self.Topmost = True  # Keep on top

        self.create_controls()
        self.create_layout()

    def load_color_map(self):
        self.color_map = read_color_map_from_excel(self.excel_path)
        if not self.color_map:
            forms.MessageBox.Show("Error loading Excel file", "Error")
            self.Close(False)

    def create_controls(self):
        self.key_label = forms.Label()
        self.key_label.Text = "Select Metadata Key:"

        self.key_dropdown = forms.DropDown()
        available_keys = sorted(self.color_map.keys())
        for key in available_keys:
            self.key_dropdown.Items.Add(key)
        if available_keys:
            self.key_dropdown.SelectedIndex = 0
            self.active_key = available_keys[0]
        self.key_dropdown.SelectedIndexChanged += self.on_key_changed

        self.default_color_label = forms.Label()
        self.default_color_label.Text = "Default Color:"

        self.default_color_picker = forms.ColorPicker()
        self.default_color_picker.Value = drawing.Color.FromArgb(128, 128, 128)
        self.default_color_picker.ValueChanged += self.on_default_color_changed

        self.legend_label = forms.Label()
        self.legend_label.Text = "Color Legend:"

        self.legend_panel = forms.Panel()
        self.legend_panel.Size = drawing.Size(450, 200)
        self.legend_panel.BackgroundColor = drawing.Colors.White
        self.update_legend()

        self.warnings_label = forms.Label()
        self.warnings_label.Text = "Warnings & Errors:"

        self.warnings_text = forms.TextArea()
        self.warnings_text.ReadOnly = True
        self.warnings_text.Size = drawing.Size(450, 100)
        self.warnings_text.Text = "Click 'Update Colors' to analyze..."

        self.update_button = forms.Button()
        self.update_button.Text = "Update Colors"
        self.update_button.Click += self.on_update_clicked

        self.select_button = forms.Button()
        self.select_button.Text = "Select Problem Objects"
        self.select_button.Click += self.on_select_clicked
        self.select_button.Enabled = False

        self.close_button = forms.Button()
        self.close_button.Text = "Close"
        self.close_button.Click += self.on_close_clicked

        self.status_label = forms.Label()
        self.status_label.Text = f"Ready - {len(self.objects)} objects"

    def create_layout(self):
        layout = forms.DynamicLayout()
        layout.Spacing = drawing.Size(5, 5)
        layout.AddRow(self.key_label)
        layout.AddRow(self.key_dropdown)
        layout.AddRow(None)
        layout.AddRow(self.default_color_label)
        layout.AddRow(self.default_color_picker)
        layout.AddRow(None)
        layout.AddRow(self.legend_label)
        layout.AddRow(self.legend_panel)
        layout.AddRow(None)
        layout.AddRow(self.warnings_label)
        layout.AddRow(self.warnings_text)
        layout.AddRow(None)

        button_layout = forms.DynamicLayout()
        button_layout.Spacing = drawing.Size(5, 5)
        button_layout.AddRow(self.update_button, self.select_button, self.close_button)
        layout.AddRow(button_layout)
        layout.AddRow(self.status_label)

        self.Content = layout

    def update_legend(self):
        if not self.active_key:
            return
        legend_layout = forms.DynamicLayout()
        legend_layout.Spacing = drawing.Size(5, 5)
        legend_layout.Padding = drawing.Padding(10)

        title = forms.Label()
        title.Text = self.active_key
        title.Font = drawing.Font(drawing.SystemFont.Bold, 12)
        legend_layout.AddRow(title)
        legend_layout.AddRow(None)

        key_colors = self.color_map.get(self.active_key, {})
        for value in sorted(key_colors.keys()):
            color_rgb = key_colors[value]

            color_panel = forms.Panel()
            color_panel.Size = drawing.Size(30, 20)
            color_panel.BackgroundColor = drawing.Color.FromArgb(color_rgb[0], color_rgb[1], color_rgb[2])

            value_label = forms.Label()
            value_label.Text = value

            rgb_text = f"RGB: {color_rgb[0]},{color_rgb[1]},{color_rgb[2]}"
            rgb_label = forms.Label()
            rgb_label.Text = rgb_text
            rgb_label.Font = drawing.Font(drawing.SystemFont.Default, 9)
            rgb_label.TextColor = drawing.Colors.Gray

            row_layout = forms.DynamicLayout()
            row_layout.Spacing = drawing.Size(5, 5)
            row_layout.AddRow(color_panel, value_label, rgb_label)
            legend_layout.AddRow(row_layout)

        self.legend_panel.Content = legend_layout

    def update_warnings(self, stats):
        warnings = []
        self.problem_objects = []

        if stats['missing_color_definition']:
            warnings.append("ERROR: Values in objects but missing from Excel:")
            warnings.append("=" * 50)
            for value, obj_list in stats['missing_color_definition']:
                warnings.append(f"  '{value}' ({len(obj_list)} objects)")
                self.problem_objects.extend(obj_list)
            warnings.append("")

        if stats['unused_color_definitions']:
            warnings.append("WARNING: Values in Excel but not in selection:")
            warnings.append("=" * 50)
            for value in stats['unused_color_definitions']:
                warnings.append(f"  '{value}'")
            warnings.append("")

        if stats['default'] > 0:
            warnings.append(f"INFO: {stats['default']} objects without '{self.active_key}' key")
            warnings.append("")

        if not warnings:
            warnings.append("SUCCESS: All objects colored correctly!")
            warnings.append(f"  {stats['colored']} objects colored")

        self.warnings_text.Text = "\n".join(warnings)
        self.select_button.Enabled = len(self.problem_objects) > 0

    def on_key_changed(self, sender, e):
        selected = self.key_dropdown.SelectedValue
        if selected:
            self.active_key = str(selected)
            self.update_legend()
            self.warnings_text.Text = "Click 'Update Colors' to analyze..."
            self.select_button.Enabled = False

    def on_default_color_changed(self, sender, e):
        color = self.default_color_picker.Value
        self.default_color = (color.R, color.G, color.B)

    def on_update_clicked(self, sender, e):
        if not self.active_key:
            return

        self.status_label.Text = "Applying colors..."
        self.load_color_map()
        self.update_legend()

        stats = apply_colors_to_objects(self.objects, self.active_key, self.color_map, self.default_color)
        sc.doc.Views.Redraw()

        self.update_warnings(stats)
        self.status_label.Text = f"Updated! Colored: {stats['colored']}, Default: {stats['default']}"

    def on_select_clicked(self, sender, e):
        if self.problem_objects:
            rs.UnselectAllObjects()
            rs.SelectObjects(self.problem_objects)
            forms.MessageBox.Show(f"Selected {len(self.problem_objects)} objects with missing colors", "Objects Selected")

    def on_close_clicked(self, sender, e):
        self.Close(True)


# ============================================================================
# MAIN UNIFIED INTERFACE
# ============================================================================

class DataVisualizationTool(forms.Dialog[bool]):
    """Main unified interface - stays on top"""

    def __init__(self):
        # Initialize base Dialog class first
        super(DataVisualizationTool, self).__init__()

        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_folder = os.path.join(self.script_dir, "_ExcelOutput")

        # Create output folder if it doesn't exist
        if not os.path.exists(self.excel_folder):
            try:
                os.makedirs(self.excel_folder)
            except:
                pass

        self.Title = "Data Visualization Tool (Excel)"
        self.Padding = drawing.Padding(15)
        self.Resizable = False
        self.Size = drawing.Size(450, 420)
        self.Topmost = True  # KEEP ON TOP

        self.create_controls()
        self.create_layout()

    def create_controls(self):
        self.header = forms.Label()
        self.header.Text = "Rhino Metadata Visualization"
        self.header.Font = drawing.Font(drawing.SystemFont.Bold, 12)

        self.step1_label = forms.Label()
        self.step1_label.Text = "Step 1: Initialize Keys"
        self.step1_label.Font = drawing.Font(drawing.SystemFont.Bold, 10)

        self.step1_desc = forms.Label()
        self.step1_desc.Text = "Apply metadata keys from Excel to objects"
        self.step1_desc.TextColor = drawing.Colors.Gray

        self.step1_button = forms.Button()
        self.step1_button.Text = "Select Excel & Apply Keys"
        self.step1_button.Click += self.on_step1

        self.step2_label = forms.Label()
        self.step2_label.Text = "Step 2: Extract Unique Values"
        self.step2_label.Font = drawing.Font(drawing.SystemFont.Bold, 10)

        self.step2_desc = forms.Label()
        self.step2_desc.Text = "Scan objects and generate color Excel"
        self.step2_desc.TextColor = drawing.Colors.Gray

        self.step2_button = forms.Button()
        self.step2_button.Text = "Scan & Generate Excel"
        self.step2_button.Click += self.on_step2

        self.step3_label = forms.Label()
        self.step3_label.Text = "Step 3: Visualize with Colors"
        self.step3_label.Font = drawing.Font(drawing.SystemFont.Bold, 10)

        self.step3_desc = forms.Label()
        self.step3_desc.Text = "Interactive color visualization"
        self.step3_desc.TextColor = drawing.Colors.Gray

        self.step3_button = forms.Button()
        self.step3_button.Text = "Open Color Manager"
        self.step3_button.Click += self.on_step3

        self.status = forms.Label()
        self.status.Text = "Ready"
        self.status.TextColor = drawing.Colors.Blue

        self.close_button = forms.Button()
        self.close_button.Text = "Close"
        self.close_button.Click += self.on_close

    def create_layout(self):
        layout = forms.DynamicLayout()
        layout.Spacing = drawing.Size(8, 8)
        layout.AddRow(self.header)
        layout.AddRow(None)
        layout.AddRow(self.step1_label)
        layout.AddRow(self.step1_desc)
        layout.AddRow(self.step1_button)
        layout.AddRow(None)
        layout.AddRow(self.step2_label)
        layout.AddRow(self.step2_desc)
        layout.AddRow(self.step2_button)
        layout.AddRow(None)
        layout.AddRow(self.step3_label)
        layout.AddRow(self.step3_desc)
        layout.AddRow(self.step3_button)
        layout.AddRow(None)
        layout.AddRow(self.status)
        layout.AddRow(self.close_button)
        self.Content = layout

    def on_step1(self, sender, e):
        excel_path = rs.OpenFileName("Select Excel File", "Excel Files (*.xlsx)|*.xlsx||",
                                     folder=self.excel_folder if os.path.exists(self.excel_folder) else None)
        if not excel_path:
            return
        objects = rs.GetObjects("Select objects", preselect=True)
        if not objects:
            return
        keys = read_keys_from_excel(excel_path)
        if not keys:
            forms.MessageBox.Show("No keys found in Excel file", "Error")
            return
        stats = apply_keys_to_objects(objects, keys)
        self.status.Text = f"Step 1: {stats['keys_added']} keys added"
        forms.MessageBox.Show(f"Keys applied!\n\nObjects: {stats['objects_processed']}\nKeys added: {stats['keys_added']}\nSkipped: {stats['keys_skipped']}", "Complete")
        self.BringToFront()  # Return focus

    def on_step2(self, sender, e):
        objects = rs.GetObjects("Select objects", preselect=True)
        if not objects:
            return
        kvd = collect_unique_values(objects)
        if not kvd:
            forms.MessageBox.Show("No metadata found", "Error")
            return
        excel_path = rs.SaveFileName("Save Excel File", "Excel Files (*.xlsx)|*.xlsx||",
                                     folder=self.excel_folder if os.path.exists(self.excel_folder) else None,
                                     filename="UniqueValuesColorSettings.xlsx")
        if not excel_path:
            return
        rows = write_color_map_excel(excel_path, kvd)
        self.status.Text = f"Step 2: {rows} values exported"
        forms.MessageBox.Show(f"Excel generated!\n\nKeys: {len(kvd)}\nValues: {rows}\n\nNext: Fill RGB colors in Excel", "Complete")
        self.BringToFront()  # Return focus

    def on_step3(self, sender, e):
        objects = rs.GetObjects("Select objects", preselect=True)
        if not objects:
            return
        excel_path = rs.OpenFileName("Select color Excel file", "Excel Files (*.xlsx)|*.xlsx||",
                                     folder=self.excel_folder if os.path.exists(self.excel_folder) else None)
        if not excel_path:
            return
        # Launch Color Manager (child dialog)
        color_dlg = ColorManagerDialog(excel_path, objects)
        color_dlg.ShowModal(self)  # Modal to parent
        self.status.Text = "Step 3: Complete"
        self.BringToFront()  # Return focus

    def on_close(self, sender, e):
        self.Close(True)


def main():
    """Main entry point"""
    if not EXCEL_AVAILABLE:
        print("="*60)
        print("ERROR: openpyxl library is not installed")
        print("="*60)
        print("\nThis script requires the openpyxl library to work with Excel files.")
        print("Please install it using the following command:\n")
        print("  pip install openpyxl\n")
        print("If you're using Rhino 8's Python 3, you may need to:")
        print("  1. Open Windows Command Prompt (cmd)")
        print("  2. Navigate to Rhino's Python directory")
        print("  3. Run: python -m pip install openpyxl")
        print("="*60)
        return

    dialog = DataVisualizationTool()
    dialog.ShowModal(None)


if __name__ == "__main__":
    main()
