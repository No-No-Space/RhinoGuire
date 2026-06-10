#! python3
# -*- coding: utf-8 -*-
"""Excel reporting for TerrainTools (see DECISIONS D7).

openpyxl-only writer, mirroring the styling of ``AreaMeasurer/Lindero.py``.
``openpyxl`` is imported lazily inside :func:`write_xlsx` so that PadGrader /
WayGrader can import this module without the dependency; only CutFillReport (and
any tool that actually exports) needs it, and declares ``# r: openpyxl``.

Drawable->PNG export stays inside the tools (keeps _core free of Eto), as the
plan's open question #2 resolved.
"""


def _styles():
    from openpyxl.styles import Font, PatternFill, Alignment
    return {
        "hdr_font":  Font(bold=True, color="FFFFFF"),
        "hdr_fill":  PatternFill(fill_type="solid", fgColor="2A8B9C"),   # Mar Caribe teal
        "hdr_align": Alignment(horizontal="center", vertical="center"),
        "sec_font":  Font(bold=True, color="1F3864"),
        "sec_fill":  PatternFill(fill_type="solid", fgColor="D9E1F2"),
        "tot_font":  Font(bold=True),
        "tot_fill":  PatternFill(fill_type="solid", fgColor="F2F2F2"),
        "warn_font": Font(bold=True, color="7F6000"),
        "warn_fill": PatternFill(fill_type="solid", fgColor="FFE699"),
    }


def write_xlsx(path, summary, per_station=None, per_cell=None, meta=None):
    """Write a cut/fill report workbook.

    summary     : dict from volumes.cut_fill (KPIs).
    per_station : optional list of dicts from volumes.per_station (ways).
    per_cell    : optional list of (i, j, x, y, terrain_z, design_z, delta).
    meta        : optional dict of context rows (e.g. {'Units': 'm', ...}).
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    st = _styles()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    unit = (meta or {}).get("unit", "")
    vol_u = ("%s³" % unit) if unit else "vol"
    area_u = ("%s²" % unit) if unit else "area"

    # ── Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    r = 1
    c = ws.cell(r, 1, "TerrainTools — Cut / Fill Report")
    c.font = st["hdr_font"]
    c.fill = st["hdr_fill"]
    c.alignment = st["hdr_align"]
    ws.cell(r, 2).fill = st["hdr_fill"]
    r += 2

    if meta:
        for k, v in meta.items():
            if k == "unit":
                continue
            ws.cell(r, 1, str(k)).font = st["sec_font"]
            ws.cell(r, 2, v)
            r += 1
        r += 1

    def kpi(label, value, fmt=None, bold=False):
        nonlocal r
        lc = ws.cell(r, 1, label)
        vc = ws.cell(r, 2, value)
        if bold:
            lc.font = st["tot_font"]
            lc.fill = st["tot_fill"]
            vc.font = st["tot_font"]
            vc.fill = st["tot_fill"]
        if fmt:
            vc.number_format = fmt
        r += 1

    NUM = "#,##0.00"
    kpi("Cut volume (%s)" % vol_u,  summary["cut_volume"],  NUM)
    kpi("Fill volume (%s)" % vol_u, summary["fill_volume"], NUM)
    kpi("Net (fill - cut) (%s)" % vol_u, summary["net"], NUM, bold=True)
    r += 1
    kpi("Cut area (%s)" % area_u,  summary["cut_area"],  NUM)
    kpi("Fill area (%s)" % area_u, summary["fill_area"], NUM)
    kpi("Graded region area (%s)" % area_u, summary["region_area"], NUM)
    r += 1
    bratio = summary["balance_ratio"]
    kpi("Balance ratio (fill/cut)", ("inf" if bratio == float("inf") else bratio),
        None if bratio == float("inf") else "#,##0.000")
    kpi("Cell size (%s)" % unit, summary["cell_size"], "#,##0.###")
    kpi("Grid (nodes)", "%d x %d" % (summary["grid_nx"], summary["grid_ny"]))
    kpi("Max cut depth (%s)" % unit,  -summary["min_delta"] if summary["min_delta"] < 0 else 0.0, NUM)
    kpi("Max fill depth (%s)" % unit,  summary["max_delta"] if summary["max_delta"] > 0 else 0.0, NUM)

    # ── Depth histogram ──────────────────────────────────────────────
    hist = summary.get("depth_histogram") or []
    if hist:
        r += 1
        hc = ws.cell(r, 1, "Depth distribution")
        hc.font = st["sec_font"]
        hc.fill = st["sec_fill"]
        ws.cell(r, 2).fill = st["sec_fill"]
        r += 1
        ws.cell(r, 1, "Delta range (%s)" % unit).font = st["tot_font"]
        ws.cell(r, 2, "Cell count").font = st["tot_font"]
        r += 1
        for lo, hi, cnt in hist:
            ws.cell(r, 1, "%.3f .. %.3f" % (lo, hi))
            ws.cell(r, 2, cnt)
            r += 1

    # ── Per Station (ways) ───────────────────────────────────────────
    if per_station:
        ws2 = wb.create_sheet("Per Station")
        cols = ["Station (%s)" % unit, "Cut area (%s)" % area_u,
                "Fill area (%s)" % area_u, "Cut vol (%s)" % vol_u,
                "Fill vol (%s)" % vol_u, "Cum cut", "Cum fill", "Cum net"]
        for ci, t in enumerate(cols, 1):
            cell = ws2.cell(1, ci, t)
            cell.font = st["hdr_font"]
            cell.fill = st["hdr_fill"]
            cell.alignment = st["hdr_align"]
            ws2.column_dimensions[get_column_letter(ci)].width = 16
        rr = 2
        for s in per_station:
            ws2.cell(rr, 1, s["station"]).number_format = "#,##0.##"
            ws2.cell(rr, 2, s["cut_area"]).number_format = NUM
            ws2.cell(rr, 3, s["fill_area"]).number_format = NUM
            ws2.cell(rr, 4, s["cut_volume"]).number_format = NUM
            ws2.cell(rr, 5, s["fill_volume"]).number_format = NUM
            ws2.cell(rr, 6, s["cum_cut"]).number_format = NUM
            ws2.cell(rr, 7, s["cum_fill"]).number_format = NUM
            ws2.cell(rr, 8, s["cum_net"]).number_format = NUM
            rr += 1

    # ── Per Cell (optional) ──────────────────────────────────────────
    if per_cell:
        ws3 = wb.create_sheet("Per Cell")
        cols = ["i", "j", "X", "Y", "Terrain Z", "Design Z", "Delta"]
        for ci, t in enumerate(cols, 1):
            cell = ws3.cell(1, ci, t)
            cell.font = st["hdr_font"]
            cell.fill = st["hdr_fill"]
            cell.alignment = st["hdr_align"]
            ws3.column_dimensions[get_column_letter(ci)].width = 14
        rr = 2
        for rec in per_cell:
            for ci, val in enumerate(rec, 1):
                cell = ws3.cell(rr, ci, val)
                if ci >= 3:
                    cell.number_format = "#,##0.###"
            rr += 1

    wb.save(path)


def collect_per_cell(grade, only_region=True):
    """Flatten a GradeResult into per-cell records for the optional sheet."""
    out = []
    for j in range(grade.ny):
        for i in range(grade.nx):
            if only_region and not grade.region_mask[j][i]:
                continue
            zd = grade.z_design[j][i]
            zt = grade.z_terrain[j][i]
            if zd is None or zt is None:
                continue
            x, y = grade.node_xy(i, j)
            out.append((i, j, x, y, zt, zd, zd - zt))
    return out
