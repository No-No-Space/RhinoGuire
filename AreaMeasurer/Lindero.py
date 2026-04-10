#! python3
# r: openpyxl
# -*- coding: utf-8 -*-
# __title__ = "Lindero"
# __doc__ = """Version = 0.4
# Date    = 2026-04-10
# Author: Aquelon - aquelon@pm.me
# _____________________________________________________________________
# Description:
# Footprint area calculator for Rhino objects.
# "Footprint" = the plan area (XY projection), NOT the sum of all surfaces.
# Modeless window — Rhino stays accessible while it is open.
# _____________________________________________________________________
# Six scenarios:
#   S1 — Selected Objects:
#        Individual footprint per object, no overlap handling.
#
#   S2 — By Layer:
#        All objects on a layer. Overlapping footprints merged with a
#        Boolean Union to avoid double-counting.
#
#   S3 — By Layer Hierarchy:
#        Parent layer contains sublayers (each sublayer = one level/floor).
#        Objects carry two user text keys: an object key and a group key.
#        Overlaps removed within each sublayer.
#
#   S4 — Custom Aggregation:
#        User-defined hierarchy of attribute keys (e.g. Domain → Main Group
#        → Subgroup → Room Type). Per sublayer, footprints are merged within
#        each leaf group (same as S3). Leaf totals are then summed across
#        all sublayers. Results shown as an indented tree.
#
#   R1 — Room Analysis:
#        Aggregates individual areas by Object Key across all floors.
#        Compares totals to a Room Target Key (set in Settings).
#        Bullet chart per room type. Data source: S3 keys or S4 hierarchy
#        (configurable in Settings).
#
#   R2 — Group Analysis:
#        Same as R1 but aggregates by Group Key and compares to Group Target Key.
#        Data source: S3 keys or S4 hierarchy (configurable in Settings).
# _____________________________________________________________________
# Last update:
# - [10.04.2026] - 0.4 S4 Custom Aggregation, R1/R2 (renamed from S4/S5),
#                     R1/R2 data source toggle (S3 keys or S4 hierarchy)
# - [26.03.2026] - 0.3 Settings tab, S4/S5 bullet-chart analysis, Write Area,
#                     per-tab results areas, overlap warnings
# - [20.02.2026] - 0.2 Export results to Excel (Objects + Summary sheets)
# - [20.02.2026] - 0.1 Initial release
# _____________________________________________________________________


import rhinoscriptsyntax as rs
import Rhino
import Rhino.Geometry as rg
import scriptcontext as sc
import Eto.Drawing as drawing
import Eto.Forms as forms
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json
import System


# ══════════════════════════════════════════════════════════════════════════════
# Model helpers
# ══════════════════════════════════════════════════════════════════════════════

def get_all_user_text_keys():
    """Return sorted list of unique user text keys found on any object in the model."""
    all_objects = rs.AllObjects()
    if not all_objects:
        return []
    keys = set()
    for guid in all_objects:
        obj_keys = rs.GetUserText(guid)
        if obj_keys:
            for k in obj_keys:
                keys.add(k)
    return sorted(keys)


def all_layer_names():
    """Return sorted list of all layer full path names in the model."""
    return sorted(rs.LayerNames() or [])


def short_name(full_path):
    """Return the last segment of a layer full path (last part after '::')."""
    return full_path.split("::")[-1] if "::" in full_path else full_path


def get_layer_objects(layer_name):
    """Return GUIDs of objects directly on the given layer."""
    return rs.ObjectsByLayer(layer_name) or []


def get_child_layers(parent_layer_name):
    """Return full path names of direct child layers of the given layer."""
    prefix = parent_layer_name + "::"
    return sorted(
        name for name in (rs.LayerNames() or [])
        if name.startswith(prefix) and "::" not in name[len(prefix):]
    )


def unit_label():
    """Return the squared-unit label matching the current Rhino model unit system."""
    mapping = {
        "Millimeters": "mm²", "Centimeters": "cm²", "Meters": "m²",
        "Kilometers": "km²", "Feet": "ft²", "Inches": "in²",
    }
    return mapping.get(sc.doc.ModelUnitSystem.ToString(), "units²")


# ══════════════════════════════════════════════════════════════════════════════
# Footprint geometry
# ══════════════════════════════════════════════════════════════════════════════

def _bbox_footprint(obj_guid):
    """Bounding-box fallback: XY rectangle from the object's bounding box."""
    rhobj = sc.doc.Objects.FindId(obj_guid)
    if rhobj is None:
        return None
    bbox = rhobj.GetBoundingBox(True)
    if not bbox.IsValid:
        return None
    pts = [
        rg.Point3d(bbox.Min.X, bbox.Min.Y, 0),
        rg.Point3d(bbox.Max.X, bbox.Min.Y, 0),
        rg.Point3d(bbox.Max.X, bbox.Max.Y, 0),
        rg.Point3d(bbox.Min.X, bbox.Max.Y, 0),
        rg.Point3d(bbox.Min.X, bbox.Min.Y, 0),
    ]
    return rg.PolylineCurve(pts)


def _brep_footprint_curves(brep):
    """
    Find the bottom horizontal face(s) of a Brep and return their outer border
    curves projected to Z=0.
    Falls back to the XY bounding box if no horizontal faces are found.
    """
    tol = sc.doc.ModelAbsoluteTolerance
    proj = rg.Transform.PlanarProjection(rg.Plane.WorldXY)

    horiz = []
    for face in brep.Faces:
        u = (face.Domain(0).Min + face.Domain(0).Max) * 0.5
        v = (face.Domain(1).Min + face.Domain(1).Max) * 0.5
        normal = face.NormalAt(u, v)
        if abs(normal.Z) > 0.9:
            pt = face.PointAt(u, v)
            horiz.append((face, pt.Z))

    if not horiz:
        bbox = brep.GetBoundingBox(True)
        if not bbox.IsValid:
            return []
        pts = [
            rg.Point3d(bbox.Min.X, bbox.Min.Y, 0),
            rg.Point3d(bbox.Max.X, bbox.Min.Y, 0),
            rg.Point3d(bbox.Max.X, bbox.Max.Y, 0),
            rg.Point3d(bbox.Min.X, bbox.Max.Y, 0),
            rg.Point3d(bbox.Min.X, bbox.Min.Y, 0),
        ]
        return [rg.PolylineCurve(pts)]

    min_z = min(z for _, z in horiz)
    bottom = [f for f, z in horiz if abs(z - min_z) <= tol * 100]

    curves = []
    for face in bottom:
        if face.OuterLoop is None:
            continue
        border = face.OuterLoop.To3dCurve()
        if border is None:
            continue
        dup = border.DuplicateCurve()
        dup.Transform(proj)
        if dup.IsClosed:
            curves.append(dup)
    return curves


def get_footprint_curves(obj_guid):
    """
    Return a list of planar closed curves at Z=0 representing the footprint
    of the object. Returns an empty list if the footprint cannot be determined.
    """
    rhobj = sc.doc.Objects.FindId(obj_guid)
    if rhobj is None:
        return []
    geom = rhobj.Geometry

    if isinstance(geom, (rg.Extrusion, rg.Surface)):
        brep = geom.ToBrep()
        if brep:
            return _brep_footprint_curves(brep)

    if isinstance(geom, rg.Brep):
        return _brep_footprint_curves(geom)

    if isinstance(geom, rg.Curve) and geom.IsClosed and geom.IsPlanar():
        proj = rg.Transform.PlanarProjection(rg.Plane.WorldXY)
        dup = geom.DuplicateCurve()
        dup.Transform(proj)
        return [dup]

    fallback = _bbox_footprint(obj_guid)
    return [fallback] if fallback else []


def curve_area(curve):
    """Return the area enclosed by a planar closed curve, or 0 on failure."""
    try:
        amp = rg.AreaMassProperties.Compute(curve)
        return amp.Area if amp else 0.0
    except Exception:
        return 0.0


def get_footprint_area(obj_guid):
    """Footprint area of a single object (sum of bottom face areas)."""
    return sum(curve_area(c) for c in get_footprint_curves(obj_guid))


def combined_area(obj_guids):
    """
    Total footprint area for a list of objects after Boolean Union on their
    projected outlines (removes overlapping footprints).
    Returns (area: float, union_succeeded: bool).
    """
    all_curves = []
    for guid in obj_guids:
        all_curves.extend(get_footprint_curves(guid))

    if not all_curves:
        return 0.0, False
    if len(all_curves) == 1:
        return curve_area(all_curves[0]), True

    tol = sc.doc.ModelAbsoluteTolerance
    try:
        unioned = rg.Curve.CreateBooleanUnion(all_curves, tol)
    except Exception:
        unioned = None

    if unioned and len(unioned) > 0:
        return sum(curve_area(c) for c in unioned), True

    return sum(curve_area(c) for c in all_curves), False


# ══════════════════════════════════════════════════════════════════════════════
# Calculation engines
# ══════════════════════════════════════════════════════════════════════════════

def _label(guid, key):
    """Object display label: user text key → Rhino object name → short GUID."""
    if key:
        val = rs.GetUserText(guid, key)
        if val:
            return val
    name = rs.ObjectName(guid)
    return name if name else str(guid)[:8] + "…"


def calc_s1(name_key):
    """Scenario 1 — Selected objects. Returns list of {guid, name, area}."""
    guids = rs.SelectedObjects() or []
    return [{"guid": str(g), "name": _label(g, name_key), "area": get_footprint_area(g)} for g in guids]


def calc_s2(layer_name, obj_key):
    """
    Scenario 2 — All objects on one layer, footprints merged.
    Returns {objects, total, union_ok}.
    """
    guids = get_layer_objects(layer_name)
    per_obj = [{"guid": str(g), "name": _label(g, obj_key), "area": get_footprint_area(g)} for g in guids]
    total, union_ok = combined_area(guids)
    return {"objects": per_obj, "total": total, "union_ok": union_ok}


def calc_s3(parent_layer, obj_key, grp_key):
    """
    Scenario 3 — Sublayer hierarchy. Overlaps removed per sublayer.
    Returns {sublayers: {name: {objects, group_totals, total, union_ok}}, overall_total}.
    """
    result = {"sublayers": {}, "overall_total": 0.0}
    for sl in get_child_layers(parent_layer):
        guids = get_layer_objects(sl)
        if not guids:
            result["sublayers"][sl] = {
                "objects": [], "group_totals": {}, "total": 0.0, "union_ok": True
            }
            continue

        objects = []
        groups = {}
        for guid in guids:
            grp_val = (rs.GetUserText(guid, grp_key) if grp_key else None) or "—"
            objects.append({
                "guid": str(guid),
                "name": _label(guid, obj_key),
                "group": grp_val,
                "area": get_footprint_area(guid),
            })
            groups.setdefault(grp_val, []).append(guid)

        total, union_ok = combined_area(guids)
        group_totals = {gv: combined_area(gguids)[0] for gv, gguids in groups.items()}

        result["sublayers"][sl] = {
            "objects": objects,
            "group_totals": group_totals,
            "total": total,
            "union_ok": union_ok,
        }
        result["overall_total"] += total

    return result


def calc_s4(parent_layer, key_sequence):
    """
    Scenario 4 — Custom Aggregation.
    key_sequence: ordered list of user text key names.
    Per sublayer, objects are grouped by their full key-value path tuple.
    Footprints within each leaf group are merged (Boolean Union) per sublayer,
    matching S3 overlap handling. Leaf totals are then summed across sublayers.
    Returns {tree, overall_total, warnings}.

    tree: nested dict  {value_str: {"area": float, "children": {...}}}
    Each node's "area" = cumulative sum of all descendant leaf areas.
    """
    if not key_sequence:
        return {"tree": {}, "overall_total": 0.0, "warnings": ["[!] No keys defined."]}

    # flat_buckets[path_tuple] = cumulative area across all floors
    flat_buckets = {}

    for sl in get_child_layers(parent_layer):
        guids = get_layer_objects(sl)
        if not guids:
            continue

        # Group objects by their full key path
        path_groups = {}
        for guid in guids:
            path = tuple(
                (rs.GetUserText(guid, k) if k else None) or "—"
                for k in key_sequence
            )
            path_groups.setdefault(path, []).append(guid)

        # Per leaf group in this floor: merged footprint (Boolean Union)
        for path, path_guids in path_groups.items():
            area, _ = combined_area(path_guids)
            flat_buckets[path] = flat_buckets.get(path, 0.0) + area

    # Build nested tree — each ancestor accumulates all descendant leaf areas
    tree = {}
    for path in sorted(flat_buckets):
        leaf_area = flat_buckets[path]
        node = tree
        for key_val in path:
            if key_val not in node:
                node[key_val] = {"area": 0.0, "children": {}}
            node[key_val]["area"] += leaf_area
            node = node[key_val]["children"]

    overall_total = sum(v["area"] for v in tree.values())

    warnings = []
    if not flat_buckets:
        warnings.append("[!] No objects found in any sublayer.")

    return {"tree": tree, "overall_total": overall_total, "warnings": warnings}


def calc_r1(parent_layer, obj_key, room_target_key):
    """
    R1 — Room Analysis.
    Sums individual footprint areas across all sublayers, grouped by obj_key value.
    Compares each group total to room_target_key.
    Returns {entries: [{label, measured, goal}], warnings: [str]}.
    """
    buckets = {}
    for sl in get_child_layers(parent_layer):
        for guid in (get_layer_objects(sl) or []):
            nv = (rs.GetUserText(guid, obj_key) if obj_key else None)
            if not nv:
                nv = rs.ObjectName(guid) or (str(guid)[:8] + "…")
            area = get_footprint_area(guid)
            if nv not in buckets:
                buckets[nv] = {"area": area, "guids": [guid]}
            else:
                buckets[nv]["area"] += area
                buckets[nv]["guids"].append(guid)

    entries, warnings = [], []
    for nv, bkt in sorted(buckets.items()):
        goal = None
        if room_target_key:
            for guid in bkt["guids"]:
                raw = rs.GetUserText(guid, room_target_key)
                if raw:
                    try:
                        goal = float(raw)
                        break
                    except (ValueError, TypeError):
                        pass
            if goal is None:
                warnings.append(f"[!] Room Target Key '{room_target_key}' not found on '{nv}'")
        entries.append({"label": nv, "measured": bkt["area"], "goal": goal})

    return {"entries": entries, "warnings": warnings}


def calc_r2(parent_layer, grp_key, grp_target_key):
    """
    R2 — Group Analysis.
    Sums individual footprint areas across all sublayers, grouped by grp_key value.
    Compares each group total to grp_target_key.
    Returns {entries: [{label, measured, goal}], warnings: [str]}.
    """
    buckets = {}
    for sl in get_child_layers(parent_layer):
        for guid in (get_layer_objects(sl) or []):
            nv = (rs.GetUserText(guid, grp_key) if grp_key else None) or "—"
            area = get_footprint_area(guid)
            if nv not in buckets:
                buckets[nv] = {"area": area, "guids": [guid]}
            else:
                buckets[nv]["area"] += area
                buckets[nv]["guids"].append(guid)

    entries, warnings = [], []
    for nv, bkt in sorted(buckets.items()):
        goal = None
        if grp_target_key:
            for guid in bkt["guids"]:
                raw = rs.GetUserText(guid, grp_target_key)
                if raw:
                    try:
                        goal = float(raw)
                        break
                    except (ValueError, TypeError):
                        pass
            if goal is None:
                warnings.append(f"[!] Group Target Key '{grp_target_key}' not found on '{nv}'")
        entries.append({"label": nv, "measured": bkt["area"], "goal": goal})

    return {"entries": entries, "warnings": warnings}


# ══════════════════════════════════════════════════════════════════════════════
# Results formatting (monospace text for the TextArea)
# ══════════════════════════════════════════════════════════════════════════════

W = 62  # result panel width in characters


def _rule(char="═"):
    return char * W


def _row(left, right="", lw=38):
    return f"  {left:<{lw}}{right:>{W - lw - 2}}"


def _fmt(v):
    return f"{v:,.4f}"


def format_s1(results, unit):
    if not results:
        return "  No objects selected.\n"
    lines = [
        _rule(),
        f"  SCENARIO 1 — SELECTED OBJECTS   [{unit}]",
        _rule("─"),
        _row("Object", "Area"),
        _rule("─"),
    ]
    for r in results:
        lines.append(_row(r["name"][:36], _fmt(r["area"])))
    total = sum(r["area"] for r in results)
    lines += [_rule("─"), _row("TOTAL", _fmt(total)), _rule(), ""]
    return "\n".join(lines)


def format_s2(data, layer_name, obj_key, unit):
    if not data["objects"]:
        return f"  No objects found on layer '{layer_name}'.\n"
    raw_sum = sum(o["area"] for o in data["objects"])
    union_note = "" if data["union_ok"] else "  [union failed — sum shown]"
    lines = [
        _rule(),
        f"  SCENARIO 2 — BY LAYER   [{unit}]",
        f"  Layer : {layer_name}",
        f"  Key   : {obj_key or '(object name / GUID)'}",
        _rule("─"),
        _row("Object", "Area"),
        _rule("─"),
    ]
    for o in data["objects"]:
        lines.append(_row(o["name"][:36], _fmt(o["area"])))
    overlap = raw_sum - data["total"]
    lines += [
        _rule("─"),
        _row("Sum (individual totals)", _fmt(raw_sum)),
        _row(f"TOTAL (footprint, overlaps merged){union_note}", _fmt(data["total"])),
    ]
    if overlap > 1e-6:
        lines += [
            _rule("─"),
            _row("  [!] Overlapping area (sum - total)", _fmt(overlap)),
            "      Some objects share footprint area.",
            "      Verify whether double-counting is intentional.",
        ]
    lines += [_rule(), ""]
    return "\n".join(lines)


def format_s3(data, parent_layer, obj_key, grp_key, unit):
    if not data["sublayers"]:
        return f"  No sublayers found under '{parent_layer}'.\n"
    lines = [
        _rule(),
        f"  SCENARIO 3 — LAYER HIERARCHY   [{unit}]",
        f"  Parent     : {parent_layer}",
        f"  Object key : {obj_key or '(object name / GUID)'}",
        f"  Group key  : {grp_key or '—'}",
        _rule("═"),
        "",
    ]
    for sl, sl_data in data["sublayers"].items():
        sn = short_name(sl)
        union_note = "" if sl_data["union_ok"] else "  [union failed]"
        lines += [
            f"  ▸ {sn}   —   Total: {_fmt(sl_data['total'])} {unit}{union_note}",
            _rule("─"),
        ]

        if sl_data["group_totals"] and grp_key:
            lines.append(_row("Group (combined footprint)", "Area"))
            lines.append("  " + "·" * (W - 2))
            for gv, ga in sorted(sl_data["group_totals"].items()):
                lines.append(_row(f"  {gv}"[:36], _fmt(ga)))
            lines.append(_rule("─"))

        if sl_data["objects"]:
            if grp_key:
                n_w, g_w, a_w = 26, 20, W - 26 - 20 - 4
                lines.append(f"  {'Object':<{n_w}}{'Group':<{g_w}}{'Area':>{a_w}}")
                lines.append("  " + "·" * (W - 2))
                for o in sl_data["objects"]:
                    n = o["name"][:n_w - 1]
                    g = o["group"][:g_w - 1]
                    a = _fmt(o["area"])
                    lines.append(f"  {n:<{n_w}}{g:<{g_w}}{a:>{a_w}}")
            else:
                lines.append(_row("Object", "Area"))
                lines.append("  " + "·" * (W - 2))
                for o in sl_data["objects"]:
                    lines.append(_row(o["name"][:36], _fmt(o["area"])))

        # Overlap warning
        individual_sum = sum(o["area"] for o in sl_data["objects"])
        group_sum = sum(sl_data["group_totals"].values()) if sl_data["group_totals"] else individual_sum
        total_overlap = individual_sum - sl_data["total"]
        cross_group_overlap = group_sum - sl_data["total"]
        if total_overlap > 1e-6:
            lines += [
                _rule("─"),
                _row("  [!] Overlapping area (sum - total)", _fmt(total_overlap)),
            ]
            if grp_key and cross_group_overlap > 1e-6:
                lines.append(_row("      of which cross-group overlap", _fmt(cross_group_overlap)))
            lines.append("      Some objects share footprint area.")
            lines.append("      Verify whether double-counting is intentional.")
        lines.append("")

    lines += [
        _rule("═"),
        _row("OVERALL TOTAL (sum of all levels)", _fmt(data["overall_total"])),
        _rule(),
        "",
    ]
    return "\n".join(lines)


def format_s4(data, parent_layer, key_sequence, unit):
    """Format the S4 custom aggregation tree as indented monospace text."""
    if not data["tree"]:
        return f"  No data found under '{parent_layer}'.\n"

    key_path_str = " > ".join(key_sequence) if key_sequence else "—"
    lines = [
        _rule(),
        f"  SCENARIO 4 — CUSTOM AGGREGATION   [{unit}]",
        f"  Parent : {parent_layer}",
        f"  Keys   : {key_path_str}",
        _rule("─"),
    ]

    def walk(node, level):
        for val in sorted(node):
            entry = node[val]
            indent = "    " * level
            if entry["children"]:
                label = f"  {indent}▸ {val}"
            else:
                label = f"  {indent}  {val}"
            lines.append(_row(label[:38], _fmt(entry["area"])))
            if entry["children"]:
                walk(entry["children"], level + 1)

    walk(data["tree"], 0)

    lines += [_rule("─"), _row("OVERALL TOTAL", _fmt(data["overall_total"])), _rule(), ""]

    if data.get("warnings"):
        lines.insert(-1, "\n".join(f"  {w}" for w in data["warnings"]))

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
# Bullet chart drawing helpers (R1 / R2)
# ══════════════════════════════════════════════════════════════════════════════

_CHART_ROW_H   = 54
_CHART_LABEL_W = 128
_CHART_VALUE_W = 108
_CHART_BAR_H   = 18


def _rgb(r, g, b):
    """Create an Eto.Drawing.Color from 0-255 integer RGB values."""
    return drawing.Color(r / 255.0, g / 255.0, b / 255.0)


def _draw_bullet_row(g, i, entry, tol, unit, total_w):
    """Draw one bullet-chart row at vertical slot i."""
    y0    = i * _CHART_ROW_H + 4
    label = entry["label"]
    meas  = entry["measured"]
    goal  = entry.get("goal")

    font    = drawing.Font("Arial", 8)
    sm_font = drawing.Font("Arial", 7)
    c_black = _rgb(0, 0, 0)

    # Left label (truncated)
    g.DrawText(font, c_black,
               drawing.PointF(4.0, float(y0 + (_CHART_ROW_H - 12) // 2)),
               label[:17])

    chart_x = float(_CHART_LABEL_W)
    chart_w = float(total_w - _CHART_LABEL_W - _CHART_VALUE_W)
    bar_y   = float(y0 + (_CHART_ROW_H - _CHART_BAR_H) // 2)
    bar_h   = float(_CHART_BAR_H)

    if chart_w < 10:
        return

    if goal is None or goal <= 0:
        g.DrawText(font, _rgb(160, 60, 60),
                   drawing.PointF(chart_x + 4.0, bar_y),
                   f"{meas:,.2f} (no target)")
        return

    max_val = max(goal * 1.35, meas * 1.05) if meas > goal else goal * 1.35
    if max_val <= 0:
        return

    def px(v):
        return chart_x + float(v) / max_val * chart_w

    # 1. Grey background
    g.FillRectangle(_rgb(215, 215, 215),
                    drawing.RectangleF(chart_x, bar_y, chart_w, bar_h))

    # 2. Yellow zone: goal*(1-tol) → goal
    yl = px(max(0.0, goal * (1.0 - tol)))
    yr = px(goal)
    if yr > yl:
        g.FillRectangle(_rgb(255, 235, 100),
                        drawing.RectangleF(yl, bar_y, yr - yl, bar_h))

    # 3. Orange zone: goal → goal*(1+tol)
    ol  = px(goal)
    or_ = px(min(max_val, goal * (1.0 + tol)))
    if or_ > ol:
        g.FillRectangle(_rgb(255, 180, 70),
                        drawing.RectangleF(ol, bar_y, or_ - ol, bar_h))

    # 4. Measured bar
    mx = px(meas) - chart_x
    if mx > 0:
        g.FillRectangle(_rgb(70, 100, 145),
                        drawing.RectangleF(chart_x, bar_y,
                                           min(float(mx), chart_w), bar_h))

    # 5. Goal line (2 px)
    gx = px(goal)
    g.DrawLine(drawing.Pen(_rgb(30, 30, 30), 2.0),
               drawing.PointF(gx, bar_y - 2.0),
               drawing.PointF(gx, bar_y + bar_h + 2.0))

    # 6. Tolerance marker lines
    tpen = drawing.Pen(_rgb(120, 120, 120), 1.0)
    for tx in (px(goal * (1.0 - tol)), px(goal * (1.0 + tol))):
        if chart_x <= tx <= chart_x + chart_w:
            g.DrawLine(tpen,
                       drawing.PointF(tx, bar_y),
                       drawing.PointF(tx, bar_y + bar_h))

    # 7. Right labels: measured/goal and delta %
    delta = (meas - goal) / goal * 100.0
    sign  = "+" if delta >= 0 else ""
    vx    = float(total_w - _CHART_VALUE_W + 4)
    g.DrawText(font,    c_black,
               drawing.PointF(vx, bar_y - 1.0),
               f"{meas:,.1f}/{goal:,.1f}")
    g.DrawText(sm_font, _rgb(80, 80, 80),
               drawing.PointF(vx, bar_y + 11.0),
               f"{sign}{delta:.1f}%  [{unit}]")


def _export_chart_png(entries, tol, unit, path, chart_width=900):
    """Render bullet-chart entries to a PNG file at the given path."""
    n = max(1, len(entries))
    height = n * _CHART_ROW_H + 20
    bmp = drawing.Bitmap(chart_width, height, drawing.PixelFormat.Format32bppRgba)
    g = drawing.Graphics(bmp)
    try:
        g.FillRectangle(drawing.Colors.White,
                        drawing.RectangleF(0.0, 0.0, float(chart_width), float(height)))
        for i, entry in enumerate(entries):
            _draw_bullet_row(g, i, entry, tol, unit, chart_width)
    finally:
        g.Dispose()
    bmp.Save(path)
    bmp.Dispose()


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

class LinderoForm(forms.Form):
    """Modeless footprint area calculator — stays open between calculations."""

    def __init__(self):
        super().__init__()
        self.Title = "Lindero — Footprint Area Calculator"
        self.Padding = drawing.Padding(10)
        self.Resizable = True
        self.MinimumSize = drawing.Size(480, 540)
        self.ClientSize  = drawing.Size(680, 720)
        self.Owner = Rhino.UI.RhinoEtoApp.MainWindow

        self.available_keys   = get_all_user_text_keys()
        self.available_layers = all_layer_names()

        # Last-calculation state (for Write Area and Export)
        self._export_data = None
        self._last_s1 = None   # list of {guid, name, area}
        self._last_s2 = None   # {objects, total, union_ok}
        self._last_s3 = None   # {sublayers, overall_total}
        self._last_s4 = None   # {tree, overall_total, warnings}

        # R1 / R2 chart state
        self._r1_entries = []
        self._r1_tol     = 0.10
        self._r1_unit    = ""
        self._r2_entries = []
        self._r2_tol     = 0.10
        self._r2_unit    = ""

        self._build_ui()

    # ------------------------------------------------------------------
    # Layout builders
    # ------------------------------------------------------------------

    def _build_ui(self):
        outer = forms.StackLayout()
        outer.Orientation = forms.Orientation.Vertical
        outer.Spacing = 8
        outer.HorizontalContentAlignment = forms.HorizontalAlignment.Stretch

        # Tabs — expand to fill available height
        self.tabs = forms.TabControl()
        self.tabs.Pages.Add(self._tab_s1())
        self.tabs.Pages.Add(self._tab_s2())
        self.tabs.Pages.Add(self._tab_s3())
        self.tabs.Pages.Add(self._tab_s4())
        self.tabs.Pages.Add(self._tab_r1())
        self.tabs.Pages.Add(self._tab_r2())
        self.tabs.Pages.Add(self._tab_settings())
        outer.Items.Add(forms.StackLayoutItem(self.tabs, True))

        # Button row
        btn_row = forms.StackLayout()
        btn_row.Orientation = forms.Orientation.Horizontal
        btn_row.Spacing = 6

        calc_btn = forms.Button()
        calc_btn.Text = "Calculate"
        calc_btn.Click += self.on_calculate

        clear_btn = forms.Button()
        clear_btn.Text = "Clear"
        clear_btn.Click += self.on_clear

        refresh_btn = forms.Button()
        refresh_btn.Text = "Refresh Model"
        refresh_btn.Click += self.on_refresh_model

        export_btn = forms.Button()
        export_btn.Text = "Export to Excel"
        export_btn.Click += self.on_export

        write_btn = forms.Button()
        write_btn.Text = "Write Area to Objects"
        write_btn.Click += self.on_write_area_toggle

        png_btn = forms.Button()
        png_btn.Text = "Export Chart as PNG"
        png_btn.Click += self.on_export_png

        btn_row.Items.Add(forms.StackLayoutItem(calc_btn))
        btn_row.Items.Add(forms.StackLayoutItem(clear_btn))
        btn_row.Items.Add(forms.StackLayoutItem(refresh_btn))
        btn_row.Items.Add(forms.StackLayoutItem(export_btn))
        btn_row.Items.Add(forms.StackLayoutItem(write_btn))
        btn_row.Items.Add(forms.StackLayoutItem(png_btn))
        outer.Items.Add(forms.StackLayoutItem(btn_row))

        # Write panel (hidden until toggled)
        outer.Items.Add(forms.StackLayoutItem(self._build_write_panel()))

        # Status label
        self.status_label = forms.Label()
        self.status_label.Text = (
            f"Ready  —  {len(self.available_layers)} layer(s), "
            f"{len(self.available_keys)} key(s) loaded."
        )
        self.status_label.TextColor = drawing.Colors.Gray
        outer.Items.Add(forms.StackLayoutItem(self.status_label))

        self.Content = outer

    def _build_write_panel(self):
        self._write_panel = forms.StackLayout()
        self._write_panel.Orientation = forms.Orientation.Horizontal
        self._write_panel.Spacing = 6
        self._write_panel.Padding = drawing.Padding(0, 2, 0, 2)
        self._write_panel.Visible = False

        wk_lbl = forms.Label()
        wk_lbl.Text = "Write to Key:"

        self._write_key_combo = forms.ComboBox()
        self._write_key_combo.DataStore = ["Area"] + list(self.available_keys)
        self._write_key_combo.Text = "Area"
        self._write_key_combo.Width = 180

        confirm_btn = forms.Button()
        confirm_btn.Text = "Confirm Write"
        confirm_btn.Click += self.on_confirm_write

        cancel_btn = forms.Button()
        cancel_btn.Text = "Cancel"
        cancel_btn.Click += self._on_write_cancel

        self._write_panel.Items.Add(forms.StackLayoutItem(wk_lbl))
        self._write_panel.Items.Add(forms.StackLayoutItem(self._write_key_combo))
        self._write_panel.Items.Add(forms.StackLayoutItem(confirm_btn))
        self._write_panel.Items.Add(forms.StackLayoutItem(cancel_btn))

        return self._write_panel

    # ------------------------------------------------------------------
    # Tab builders — S1 / S2 / S3
    # ------------------------------------------------------------------

    def _tab_s1(self):
        page = forms.TabPage()
        page.Text = "S1 — Selected Objects"

        controls = forms.DynamicLayout()
        controls.DefaultSpacing = drawing.Size(5, 6)
        controls.Padding = drawing.Padding(8)

        desc = forms.Label()
        desc.Text = "Individual footprint per selected object. No overlap handling."
        desc.TextColor = drawing.Colors.Gray
        controls.AddRow(desc)
        controls.AddRow(None)

        name_lbl = forms.Label()
        name_lbl.Text = "Name Key (optional):"
        self.name_key_combo = forms.ComboBox()
        self.name_key_combo.DataStore = self.available_keys
        self.name_key_combo.PlaceholderText = "Falls back to object name / GUID"
        self.name_key_combo.Width = 300
        controls.AddRow(name_lbl, self.name_key_combo)
        controls.AddRow(None)

        note = forms.Label()
        note.Text = "Select objects in Rhino (form stays open), then click Calculate."
        note.TextColor = drawing.Colors.Gray
        controls.AddRow(note)

        self.results_s1 = forms.TextArea()
        self.results_s1.ReadOnly = True
        self.results_s1.Font = drawing.Font("Courier New", 9)
        self.results_s1.Text = "Select objects and click Calculate.\n"

        layout = forms.StackLayout()
        layout.Orientation = forms.Orientation.Vertical
        layout.HorizontalContentAlignment = forms.HorizontalAlignment.Stretch
        layout.Items.Add(forms.StackLayoutItem(controls))
        layout.Items.Add(forms.StackLayoutItem(self.results_s1, True))

        page.Content = layout
        return page

    def _tab_s2(self):
        page = forms.TabPage()
        page.Text = "S2 — By Layer"

        controls = forms.DynamicLayout()
        controls.DefaultSpacing = drawing.Size(5, 6)
        controls.Padding = drawing.Padding(8)

        desc = forms.Label()
        desc.Text = "All objects on a layer. Overlapping footprints merged (Boolean Union)."
        desc.TextColor = drawing.Colors.Gray
        controls.AddRow(desc)
        controls.AddRow(None)

        layer_lbl = forms.Label()
        layer_lbl.Text = "Layer:"
        self.layer_s2_dd = forms.DropDown()
        self._populate_layer_dd(self.layer_s2_dd)
        self.layer_s2_dd.Width = 300
        controls.AddRow(layer_lbl, self.layer_s2_dd)

        obj_key_lbl = forms.Label()
        obj_key_lbl.Text = "Object Key:"
        self.obj_key_s2 = forms.ComboBox()
        self.obj_key_s2.DataStore = self.available_keys
        self.obj_key_s2.PlaceholderText = "User text key for object labels"
        self.obj_key_s2.Width = 300
        controls.AddRow(obj_key_lbl, self.obj_key_s2)

        self.results_s2 = forms.TextArea()
        self.results_s2.ReadOnly = True
        self.results_s2.Font = drawing.Font("Courier New", 9)
        self.results_s2.Text = "Select a layer and click Calculate.\n"

        layout = forms.StackLayout()
        layout.Orientation = forms.Orientation.Vertical
        layout.HorizontalContentAlignment = forms.HorizontalAlignment.Stretch
        layout.Items.Add(forms.StackLayoutItem(controls))
        layout.Items.Add(forms.StackLayoutItem(self.results_s2, True))

        page.Content = layout
        return page

    def _tab_s3(self):
        page = forms.TabPage()
        page.Text = "S3 — Layer Hierarchy"

        controls = forms.DynamicLayout()
        controls.DefaultSpacing = drawing.Size(5, 6)
        controls.Padding = drawing.Padding(8)

        desc = forms.Label()
        desc.Text = "Parent layer + sublayers (each sublayer = one level / floor)."
        desc.TextColor = drawing.Colors.Gray
        controls.AddRow(desc)
        controls.AddRow(None)

        parent_lbl = forms.Label()
        parent_lbl.Text = "Parent Layer:"
        self.parent_layer_dd = forms.DropDown()
        self._populate_layer_dd(self.parent_layer_dd)
        self.parent_layer_dd.Width = 300
        controls.AddRow(parent_lbl, self.parent_layer_dd)
        controls.AddRow(None)

        obj_key_lbl = forms.Label()
        obj_key_lbl.Text = "Object Key:"
        self.obj_key_s3 = forms.ComboBox()
        self.obj_key_s3.DataStore = self.available_keys
        self.obj_key_s3.PlaceholderText = "Individual / small group name"
        self.obj_key_s3.Width = 300
        obj_key_hint = forms.Label()
        obj_key_hint.Text = "Individual or small group name (e.g. 'Room Name')"
        obj_key_hint.TextColor = drawing.Colors.Gray
        controls.AddRow(obj_key_lbl, self.obj_key_s3)
        controls.AddRow(None, obj_key_hint)
        controls.AddRow(None)

        grp_key_lbl = forms.Label()
        grp_key_lbl.Text = "Group Key:"
        self.grp_key_s3 = forms.ComboBox()
        self.grp_key_s3.DataStore = self.available_keys
        self.grp_key_s3.PlaceholderText = "Department / class (larger grouping)"
        self.grp_key_s3.Width = 300
        grp_key_hint = forms.Label()
        grp_key_hint.Text = "Larger class grouping (e.g. 'Department')"
        grp_key_hint.TextColor = drawing.Colors.Gray
        controls.AddRow(grp_key_lbl, self.grp_key_s3)
        controls.AddRow(None, grp_key_hint)

        self.results_s3 = forms.TextArea()
        self.results_s3.ReadOnly = True
        self.results_s3.Font = drawing.Font("Courier New", 9)
        self.results_s3.Text = "Select a parent layer and click Calculate.\n"

        layout = forms.StackLayout()
        layout.Orientation = forms.Orientation.Vertical
        layout.HorizontalContentAlignment = forms.HorizontalAlignment.Stretch
        layout.Items.Add(forms.StackLayoutItem(controls))
        layout.Items.Add(forms.StackLayoutItem(self.results_s3, True))

        page.Content = layout
        return page

    # ------------------------------------------------------------------
    # Tab builder — S4 (Custom Aggregation)
    # ------------------------------------------------------------------

    def _tab_s4(self):
        page = forms.TabPage()
        page.Text = "S4 — Custom Aggregation"

        layout = forms.StackLayout()
        layout.Orientation = forms.Orientation.Vertical
        layout.Spacing = 4
        layout.Padding = drawing.Padding(8)
        layout.HorizontalContentAlignment = forms.HorizontalAlignment.Stretch

        desc = forms.Label()
        desc.Text = (
            "Groups objects by a user-defined hierarchy of attribute keys. "
            "Footprints are merged per leaf group per floor (same as S3), "
            "then summed across all floors of the chosen parent layer."
        )
        desc.TextColor = drawing.Colors.Gray
        desc.Wrap = forms.WrapMode.Word
        layout.Items.Add(forms.StackLayoutItem(desc))

        # Parent layer row
        pl_row = forms.DynamicLayout()
        pl_row.DefaultSpacing = drawing.Size(5, 4)
        pl_row.Padding = drawing.Padding(0, 4, 0, 4)
        pl_lbl = forms.Label()
        pl_lbl.Text = "Parent Layer:"
        self.parent_layer_s4_dd = forms.DropDown()
        self._populate_layer_dd(self.parent_layer_s4_dd)
        self.parent_layer_s4_dd.Width = 300
        pl_row.AddRow(pl_lbl, self.parent_layer_s4_dd)
        layout.Items.Add(forms.StackLayoutItem(pl_row))

        # Keys section header + Add button
        keys_header = forms.StackLayout()
        keys_header.Orientation = forms.Orientation.Horizontal
        keys_header.Spacing = 10
        keys_header.Padding = drawing.Padding(0, 2, 0, 2)
        keys_lbl = forms.Label()
        keys_lbl.Text = "Attribute Keys  (top row = level 1, bottom = deepest):"
        add_btn = forms.Button()
        add_btn.Text = "+ Add Level"
        add_btn.Click += self._on_s4_add_key
        keys_header.Items.Add(forms.StackLayoutItem(keys_lbl))
        keys_header.Items.Add(forms.StackLayoutItem(add_btn))
        layout.Items.Add(forms.StackLayoutItem(keys_header))

        # Dynamic key rows container
        self._s4_keys_layout = forms.StackLayout()
        self._s4_keys_layout.Orientation = forms.Orientation.Vertical
        self._s4_keys_layout.Spacing = 3
        self._s4_key_rows = []
        self._s4_add_key_row()   # seed with one empty row
        layout.Items.Add(forms.StackLayoutItem(self._s4_keys_layout))

        # Results text area
        self.results_s4 = forms.TextArea()
        self.results_s4.ReadOnly = True
        self.results_s4.Font = drawing.Font("Courier New", 9)
        self.results_s4.Text = "Define at least one key, select a parent layer, and click Calculate.\n"
        layout.Items.Add(forms.StackLayoutItem(self.results_s4, True))

        page.Content = layout
        return page

    def _s4_add_key_row(self, initial_text=""):
        """Append one attribute-key combo row to the S4 dynamic list."""
        row = forms.StackLayout()
        row.Orientation = forms.Orientation.Horizontal
        row.Spacing = 4

        cb = forms.ComboBox()
        cb.DataStore = self.available_keys
        cb.Text = initial_text
        cb.Width = 280

        rm_btn = forms.Button()
        rm_btn.Text = "✕"
        rm_btn.Width = 28

        def on_remove(s, e, r=row, c=cb):
            self._s4_remove_key_row(r, c)
        rm_btn.Click += on_remove

        row.Items.Add(forms.StackLayoutItem(cb, True))
        row.Items.Add(forms.StackLayoutItem(rm_btn))

        self._s4_key_rows.append(cb)
        self._s4_keys_layout.Items.Add(forms.StackLayoutItem(row))

    def _on_s4_add_key(self, _s, _e):
        self._s4_add_key_row()

    def _s4_remove_key_row(self, row_ctrl, cb):
        """Remove a key row, keeping at least one."""
        if len(self._s4_key_rows) <= 1:
            return
        if cb in self._s4_key_rows:
            self._s4_key_rows.remove(cb)
        for i in range(self._s4_keys_layout.Items.Count):
            if self._s4_keys_layout.Items[i].Control is row_ctrl:
                self._s4_keys_layout.Items.RemoveAt(i)
                break

    # ------------------------------------------------------------------
    # Tab builders — R1 / R2
    # ------------------------------------------------------------------

    def _tab_r1(self):
        page = forms.TabPage()
        page.Text = "R1 — Room Analysis"

        layout = forms.StackLayout()
        layout.Orientation = forms.Orientation.Vertical
        layout.Spacing = 4
        layout.Padding = drawing.Padding(8)
        layout.HorizontalContentAlignment = forms.HorizontalAlignment.Stretch

        desc = forms.Label()
        desc.Text = (
            "Aggregates individual object areas by Object Key across all floors. "
            "Compares totals to Room Target Key (set in Settings). "
            "Data source: S3 keys or S4 hierarchy (configurable in Settings)."
        )
        desc.TextColor = drawing.Colors.Gray
        desc.Wrap = forms.WrapMode.Word

        self.warn_r1 = forms.Label()
        self.warn_r1.TextColor = drawing.Color(0.6, 0.1, 0.1)
        self.warn_r1.Wrap = forms.WrapMode.Word
        self.warn_r1.Visible = False

        self.chart_r1 = forms.Drawable()
        self.chart_r1.Size = drawing.Size(400, 10)
        self.chart_r1.Paint += self._paint_r1

        scroll_r1 = forms.Scrollable()
        scroll_r1.ExpandContentWidth = True
        scroll_r1.ExpandContentHeight = False
        scroll_r1.Content = self.chart_r1

        layout.Items.Add(forms.StackLayoutItem(desc))
        layout.Items.Add(forms.StackLayoutItem(self.warn_r1))
        layout.Items.Add(forms.StackLayoutItem(scroll_r1, True))

        page.Content = layout
        return page

    def _tab_r2(self):
        page = forms.TabPage()
        page.Text = "R2 — Group Analysis"

        layout = forms.StackLayout()
        layout.Orientation = forms.Orientation.Vertical
        layout.Spacing = 4
        layout.Padding = drawing.Padding(8)
        layout.HorizontalContentAlignment = forms.HorizontalAlignment.Stretch

        desc = forms.Label()
        desc.Text = (
            "Aggregates individual object areas by Group Key across all floors. "
            "Compares totals to Group Target Key (set in Settings). "
            "Data source: S3 keys or S4 hierarchy (configurable in Settings)."
        )
        desc.TextColor = drawing.Colors.Gray
        desc.Wrap = forms.WrapMode.Word

        self.warn_r2 = forms.Label()
        self.warn_r2.TextColor = drawing.Color(0.6, 0.1, 0.1)
        self.warn_r2.Wrap = forms.WrapMode.Word
        self.warn_r2.Visible = False

        self.chart_r2 = forms.Drawable()
        self.chart_r2.Size = drawing.Size(400, 10)
        self.chart_r2.Paint += self._paint_r2

        scroll_r2 = forms.Scrollable()
        scroll_r2.ExpandContentWidth = True
        scroll_r2.ExpandContentHeight = False
        scroll_r2.Content = self.chart_r2

        layout.Items.Add(forms.StackLayoutItem(desc))
        layout.Items.Add(forms.StackLayoutItem(self.warn_r2))
        layout.Items.Add(forms.StackLayoutItem(scroll_r2, True))

        page.Content = layout
        return page

    # ------------------------------------------------------------------
    # Tab builder — Settings
    # ------------------------------------------------------------------

    def _tab_settings(self):
        page = forms.TabPage()
        page.Text = "Settings"

        layout = forms.DynamicLayout()
        layout.DefaultSpacing = drawing.Size(5, 8)
        layout.Padding = drawing.Padding(12)

        # ── Program Key Mapping ──────────────────────────────────────
        sec1 = forms.Label()
        sec1.Text = "Program Key Mapping"
        sec1.Font = drawing.Font("Arial", 9, drawing.FontStyle.Bold)
        layout.AddRow(sec1)
        layout.AddRow(None)

        rtk_lbl = forms.Label()
        rtk_lbl.Text = "Room Target Key:"
        self.room_target_key_dd = forms.ComboBox()
        self.room_target_key_dd.DataStore = self.available_keys
        self.room_target_key_dd.PlaceholderText = "Key holding target area per room"
        self.room_target_key_dd.Width = 280
        rtk_hint = forms.Label()
        rtk_hint.Text = "Key containing the target area per room type"
        rtk_hint.TextColor = drawing.Colors.Gray
        layout.AddRow(rtk_lbl, self.room_target_key_dd)
        layout.AddRow(None, rtk_hint)
        layout.AddRow(None)

        gtk_lbl = forms.Label()
        gtk_lbl.Text = "Group Target Key:"
        self.grp_target_key_dd = forms.ComboBox()
        self.grp_target_key_dd.DataStore = self.available_keys
        self.grp_target_key_dd.PlaceholderText = "Key holding target area per group"
        self.grp_target_key_dd.Width = 280
        gtk_hint = forms.Label()
        gtk_hint.Text = "Key containing the target area per group"
        gtk_hint.TextColor = drawing.Colors.Gray
        layout.AddRow(gtk_lbl, self.grp_target_key_dd)
        layout.AddRow(None, gtk_hint)
        layout.AddRow(None)

        # ── Tolerance ────────────────────────────────────────────────
        sep1 = forms.Label()
        sep1.Text = "─" * 42
        sep1.TextColor = drawing.Colors.Gray
        layout.AddRow(sep1)

        sec2 = forms.Label()
        sec2.Text = "Tolerance"
        sec2.Font = drawing.Font("Arial", 9, drawing.FontStyle.Bold)
        layout.AddRow(sec2)
        layout.AddRow(None)

        tol_lbl = forms.Label()
        tol_lbl.Text = "Global Tolerance (%):"
        self.tolerance_stepper = forms.NumericStepper()
        self.tolerance_stepper.MinValue    = 0.0
        self.tolerance_stepper.MaxValue    = 50.0
        self.tolerance_stepper.Value       = 10.0
        self.tolerance_stepper.DecimalPlaces = 1
        self.tolerance_stepper.Increment   = 0.5
        self.tolerance_stepper.Width       = 80
        tol_hint = forms.Label()
        tol_hint.Text = "Symmetric tolerance applied in R1 and R2 bullet charts"
        tol_hint.TextColor = drawing.Colors.Gray
        layout.AddRow(tol_lbl, self.tolerance_stepper)
        layout.AddRow(None, tol_hint)
        layout.AddRow(None)

        # ── R1 / R2 Data Source ───────────────────────────────────────
        sep2 = forms.Label()
        sep2.Text = "─" * 42
        sep2.TextColor = drawing.Colors.Gray
        layout.AddRow(sep2)

        sec3 = forms.Label()
        sec3.Text = "R1 / R2 Data Source"
        sec3.Font = drawing.Font("Arial", 9, drawing.FontStyle.Bold)
        layout.AddRow(sec3)
        layout.AddRow(None)

        src_lbl = forms.Label()
        src_lbl.Text = "Source:"
        self.r1r2_source_dd = forms.DropDown()
        self.r1r2_source_dd.Items.Add("S3 keys")
        self.r1r2_source_dd.Items.Add("S4 hierarchy")
        self.r1r2_source_dd.SelectedIndex = 0
        self.r1r2_source_dd.Width = 180
        src_hint = forms.Label()
        src_hint.Text = "Which parent layer and keys feed the R1 and R2 bullet charts"
        src_hint.TextColor = drawing.Colors.Gray
        layout.AddRow(src_lbl, self.r1r2_source_dd)
        layout.AddRow(None, src_hint)
        layout.AddRow(None)

        r1_lvl_lbl = forms.Label()
        r1_lvl_lbl.Text = "R1 Room Level:"
        self.r1_level_stepper = forms.NumericStepper()
        self.r1_level_stepper.MinValue     = 1
        self.r1_level_stepper.MaxValue     = 10
        self.r1_level_stepper.Value        = 1
        self.r1_level_stepper.DecimalPlaces = 0
        self.r1_level_stepper.Increment    = 1
        self.r1_level_stepper.Width        = 60
        r1_lvl_hint = forms.Label()
        r1_lvl_hint.Text = "Position in S4 key list used as room key  (1 = top level)"
        r1_lvl_hint.TextColor = drawing.Colors.Gray
        layout.AddRow(r1_lvl_lbl, self.r1_level_stepper)
        layout.AddRow(None, r1_lvl_hint)
        layout.AddRow(None)

        r2_lvl_lbl = forms.Label()
        r2_lvl_lbl.Text = "R2 Group Level:"
        self.r2_level_stepper = forms.NumericStepper()
        self.r2_level_stepper.MinValue     = 1
        self.r2_level_stepper.MaxValue     = 10
        self.r2_level_stepper.Value        = 2
        self.r2_level_stepper.DecimalPlaces = 0
        self.r2_level_stepper.Increment    = 1
        self.r2_level_stepper.Width        = 60
        r2_lvl_hint = forms.Label()
        r2_lvl_hint.Text = "Position in S4 key list used as group key  (1 = top level)"
        r2_lvl_hint.TextColor = drawing.Colors.Gray
        layout.AddRow(r2_lvl_lbl, self.r2_level_stepper)
        layout.AddRow(None, r2_lvl_hint)
        layout.AddRow(None)

        # ── Configuration ────────────────────────────────────────────
        sep3 = forms.Label()
        sep3.Text = "─" * 42
        sep3.TextColor = drawing.Colors.Gray
        layout.AddRow(sep3)

        sec4 = forms.Label()
        sec4.Text = "Configuration"
        sec4.Font = drawing.Font("Arial", 9, drawing.FontStyle.Bold)
        layout.AddRow(sec4)
        layout.AddRow(None)

        save_btn = forms.Button()
        save_btn.Text = "Save Config"
        save_btn.Click += self.on_save_config

        load_btn = forms.Button()
        load_btn.Text = "Load Config"
        load_btn.Click += self.on_load_config

        cfg_hint = forms.Label()
        cfg_hint.Text = 'Saves / loads all settings to "lindero_config.json"'
        cfg_hint.TextColor = drawing.Colors.Gray
        layout.AddRow(save_btn, load_btn)
        layout.AddRow(None, cfg_hint)

        page.Content = layout
        return page

    # ------------------------------------------------------------------
    # Layer dropdown helpers
    # ------------------------------------------------------------------

    def _populate_layer_dd(self, dd):
        dd.Items.Clear()
        for name in self.available_layers:
            dd.Items.Add(name)
        if self.available_layers:
            dd.SelectedIndex = 0

    def _selected_layer(self, dd):
        idx = dd.SelectedIndex
        if idx < 0 or idx >= len(self.available_layers):
            return None
        return self.available_layers[idx]

    def _restore_layer_dd(self, dd, prev_name):
        self._populate_layer_dd(dd)
        if prev_name and prev_name in self.available_layers:
            dd.SelectedIndex = self.available_layers.index(prev_name)

    # ------------------------------------------------------------------
    # Event handlers — model / navigation
    # ------------------------------------------------------------------

    def on_refresh_model(self, _sender, _e):
        prev_s2     = self._selected_layer(self.layer_s2_dd)
        prev_parent = self._selected_layer(self.parent_layer_dd)
        prev_s4_par = self._selected_layer(self.parent_layer_s4_dd)

        self.available_keys   = get_all_user_text_keys()
        self.available_layers = all_layer_names()

        for combo in [self.name_key_combo, self.obj_key_s2, self.obj_key_s3,
                      self.grp_key_s3, self.room_target_key_dd, self.grp_target_key_dd]:
            txt = combo.Text
            combo.DataStore = self.available_keys
            combo.Text = txt

        # Refresh S4 dynamic key rows
        for cb in self._s4_key_rows:
            txt = cb.Text
            cb.DataStore = self.available_keys
            cb.Text = txt

        self._restore_layer_dd(self.layer_s2_dd,        prev_s2)
        self._restore_layer_dd(self.parent_layer_dd,    prev_parent)
        self._restore_layer_dd(self.parent_layer_s4_dd, prev_s4_par)

        prev_wk = self._write_key_combo.Text
        self._write_key_combo.DataStore = ["Area"] + list(self.available_keys)
        self._write_key_combo.Text = prev_wk or "Area"

        self.status_label.Text = (
            f"Refreshed  —  {len(self.available_layers)} layer(s), "
            f"{len(self.available_keys)} key(s)."
        )
        self.status_label.TextColor = drawing.Colors.Gray

    def on_calculate(self, _sender, _e):
        unit = unit_label()
        try:
            idx = self.tabs.SelectedIndex
            runners = {
                0: self._run_s1,
                1: self._run_s2,
                2: self._run_s3,
                3: self._run_s4,
                4: self._run_r1,
                5: self._run_r2,
            }
            fn = runners.get(idx)
            if fn:
                fn(unit)
            else:
                self.status_label.Text = "Switch to a calculation tab (S1–S4, R1–R2) to calculate."
                self.status_label.TextColor = drawing.Colors.Gray
        except Exception as ex:
            self.status_label.Text = f"Error: {ex}"
            self.status_label.TextColor = drawing.Colors.Red

    def on_clear(self, _sender, _e):
        idx = self.tabs.SelectedIndex
        if idx == 0:
            self.results_s1.Text = ""
        elif idx == 1:
            self.results_s2.Text = ""
        elif idx == 2:
            self.results_s3.Text = ""
        elif idx == 3:
            self.results_s4.Text = ""
            self._last_s4 = None
        elif idx == 4:
            self._r1_entries = []
            self.warn_r1.Visible = False
            self.chart_r1.Invalidate()
        elif idx == 5:
            self._r2_entries = []
            self.warn_r2.Visible = False
            self.chart_r2.Invalidate()
        self.status_label.Text = "Cleared."
        self.status_label.TextColor = drawing.Colors.Gray

    def on_export(self, _sender, _e):
        if self._export_data is None:
            self.status_label.Text = "Nothing to export — run Calculate on S1, S2, S3, or S4 first."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        path = rs.SaveFileName(
            "Export Results to Excel",
            "Excel Files (*.xlsx)|*.xlsx||",
            None, "Lindero_Results", "xlsx"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            export_to_excel(self._export_data, path)
            self.status_label.Text = f"Exported → {path}"
            self.status_label.TextColor = drawing.Colors.Green
        except Exception as ex:
            self.status_label.Text = f"Export failed: {ex}"
            self.status_label.TextColor = drawing.Colors.Red

    def on_export_png(self, _sender, _e):
        idx = self.tabs.SelectedIndex
        if idx == 4:
            entries, tol, unit = self._r1_entries, self._r1_tol, self._r1_unit
            default_name = "Lindero_R1_RoomAnalysis"
        elif idx == 5:
            entries, tol, unit = self._r2_entries, self._r2_tol, self._r2_unit
            default_name = "Lindero_R2_GroupAnalysis"
        else:
            self.status_label.Text = "Switch to R1 or R2 tab to export a chart as PNG."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        if not entries:
            self.status_label.Text = "No chart data — run Calculate first."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        path = rs.SaveFileName(
            "Export Chart as PNG",
            "PNG Images (*.png)|*.png||",
            None, default_name, "png"
        )
        if not path:
            return
        if not path.lower().endswith(".png"):
            path += ".png"

        try:
            _export_chart_png(entries, tol, unit, path)
            self.status_label.Text = f"Chart exported → {path}"
            self.status_label.TextColor = drawing.Colors.Green
        except Exception as ex:
            self.status_label.Text = f"PNG export failed: {ex}"
            self.status_label.TextColor = drawing.Colors.Red

    # ------------------------------------------------------------------
    # Event handlers — Write Area
    # ------------------------------------------------------------------

    def on_write_area_toggle(self, _s, _e):
        self._write_panel.Visible = not self._write_panel.Visible

    def _on_write_cancel(self, _s, _e):
        self._write_panel.Visible = False

    def on_confirm_write(self, _s, _e):
        key = self._write_key_combo.Text.strip()
        if not key:
            self.status_label.Text = "Enter a key name to write to."
            self.status_label.TextColor = drawing.Colors.Red
            return

        idx = self.tabs.SelectedIndex
        objects = []
        if idx == 0 and self._last_s1:
            objects = self._last_s1
        elif idx == 1 and self._last_s2:
            objects = self._last_s2["objects"]
        elif idx == 2 and self._last_s3:
            for sl_data in self._last_s3["sublayers"].values():
                objects += sl_data["objects"]

        if not objects:
            self.status_label.Text = "No calculated data — run Calculate first on S1, S2, or S3."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        count = 0
        for obj in objects:
            try:
                guid   = System.Guid(obj["guid"])
                rh_obj = sc.doc.Objects.FindId(guid)
                if rh_obj:
                    rh_obj.Attributes.SetUserString(key, f"{obj['area']:.4f}")
                    rh_obj.CommitChanges()
                    count += 1
            except Exception:
                pass

        sc.doc.Views.Redraw()
        self._write_panel.Visible = False
        self.status_label.Text = f"Area written to {count} object(s) using key '{key}'"
        self.status_label.TextColor = drawing.Colors.Green

    # ------------------------------------------------------------------
    # Event handlers — Settings config
    # ------------------------------------------------------------------

    def on_save_config(self, _s, _e):
        path = rs.SaveFileName(
            "Save Lindero Configuration",
            "JSON Files (*.json)|*.json||",
            None, "lindero_config", "json"
        )
        if not path:
            return
        if not path.lower().endswith(".json"):
            path += ".json"
        cfg = {
            "room_target_key":   self.room_target_key_dd.Text.strip(),
            "group_target_key":  self.grp_target_key_dd.Text.strip(),
            "tolerance_percent": self.tolerance_stepper.Value,
            "s4_parent_layer":   self._selected_layer(self.parent_layer_s4_dd) or "",
            "s4_key_sequence":   [cb.Text.strip() for cb in self._s4_key_rows],
            "r1r2_source":       self.r1r2_source_dd.SelectedIndex,
            "r1_level_index":    int(self.r1_level_stepper.Value),
            "r2_level_index":    int(self.r2_level_stepper.Value),
        }
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(cfg, f, indent=2)
            self.status_label.Text = f"Config saved → {path}"
            self.status_label.TextColor = drawing.Colors.Green
        except Exception as ex:
            self.status_label.Text = f"Save failed: {ex}"
            self.status_label.TextColor = drawing.Colors.Red

    def on_load_config(self, _s, _e):
        path = rs.OpenFileName(
            "Load Lindero Configuration",
            "JSON Files (*.json)|*.json||"
        )
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            self.room_target_key_dd.Text  = cfg.get("room_target_key", "")
            self.grp_target_key_dd.Text   = cfg.get("group_target_key", "")
            self.tolerance_stepper.Value  = float(cfg.get("tolerance_percent", 10.0))
            self.r1r2_source_dd.SelectedIndex = int(cfg.get("r1r2_source", 0))
            self.r1_level_stepper.Value   = float(cfg.get("r1_level_index", 1))
            self.r2_level_stepper.Value   = float(cfg.get("r2_level_index", 2))

            # Restore S4 parent layer
            s4_par = cfg.get("s4_parent_layer", "")
            if s4_par and s4_par in self.available_layers:
                self.parent_layer_s4_dd.SelectedIndex = self.available_layers.index(s4_par)

            # Restore S4 key sequence
            key_seq = cfg.get("s4_key_sequence", [])
            if key_seq:
                # Clear all existing rows
                while self._s4_keys_layout.Items.Count > 0:
                    self._s4_keys_layout.Items.RemoveAt(0)
                self._s4_key_rows.clear()
                for txt in key_seq:
                    self._s4_add_key_row(txt)
                if not self._s4_key_rows:
                    self._s4_add_key_row()

            self.status_label.Text = f"Config loaded ← {path}"
            self.status_label.TextColor = drawing.Colors.Green
        except Exception as ex:
            self.status_label.Text = f"Load failed: {ex}"
            self.status_label.TextColor = drawing.Colors.Red

    # ------------------------------------------------------------------
    # Per-scenario runners — S1 / S2 / S3
    # ------------------------------------------------------------------

    def _run_s1(self, unit):
        name_key = self.name_key_combo.Text.strip()
        results  = calc_s1(name_key)

        if not results:
            self.results_s1.Text = "  No objects are currently selected in Rhino.\n"
            self.status_label.Text = "No selection."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        self.results_s1.Text = format_s1(results, unit)
        self._last_s1 = results
        total = sum(r["area"] for r in results)
        self._export_data = {
            "scenario": 1, "unit": unit,
            "params": {"name_key": name_key},
            "objects": results,
        }
        self.status_label.Text = (
            f"S1  —  {len(results)} object(s)  |  Total: {total:,.4f} {unit}"
        )
        self.status_label.TextColor = drawing.Colors.Green

    def _run_s2(self, unit):
        layer_name = self._selected_layer(self.layer_s2_dd)
        if not layer_name:
            self.status_label.Text = "Please select a layer."
            self.status_label.TextColor = drawing.Colors.Red
            return

        obj_key = self.obj_key_s2.Text.strip()
        data    = calc_s2(layer_name, obj_key)

        if not data["objects"]:
            self.results_s2.Text = f"  No objects found on layer '{layer_name}'.\n"
            self.status_label.Text = f"Layer '{short_name(layer_name)}' is empty."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        self.results_s2.Text = format_s2(data, layer_name, obj_key, unit)
        self._last_s2 = data
        self._export_data = {
            "scenario": 2, "unit": unit,
            "params": {"layer_name": layer_name, "obj_key": obj_key},
            **data,
        }
        self.status_label.Text = (
            f"S2  —  '{short_name(layer_name)}'  |  "
            f"{len(data['objects'])} object(s)  |  Total: {data['total']:,.4f} {unit}"
        )
        self.status_label.TextColor = drawing.Colors.Green

    def _run_s3(self, unit):
        parent = self._selected_layer(self.parent_layer_dd)
        if not parent:
            self.status_label.Text = "Please select a parent layer."
            self.status_label.TextColor = drawing.Colors.Red
            return

        obj_key = self.obj_key_s3.Text.strip()
        grp_key = self.grp_key_s3.Text.strip()
        data    = calc_s3(parent, obj_key, grp_key)

        non_empty = [sl for sl, d in data["sublayers"].items() if d["objects"]]
        if not non_empty:
            self.results_s3.Text = f"  No objects found in any sublayer of '{parent}'.\n"
            self.status_label.Text = "No objects found in sublayers."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        self.results_s3.Text = format_s3(data, parent, obj_key, grp_key, unit)
        self._last_s3 = data
        self._export_data = {
            "scenario": 3, "unit": unit,
            "params": {"parent": parent, "obj_key": obj_key, "grp_key": grp_key},
            "sublayers": data["sublayers"],
            "overall_total": data["overall_total"],
        }
        self.status_label.Text = (
            f"S3  —  '{short_name(parent)}'  |  "
            f"{len(data['sublayers'])} sublayer(s)  |  "
            f"Overall total: {data['overall_total']:,.4f} {unit}"
        )
        self.status_label.TextColor = drawing.Colors.Green

    # ------------------------------------------------------------------
    # Per-scenario runner — S4
    # ------------------------------------------------------------------

    def _run_s4(self, unit):
        parent = self._selected_layer(self.parent_layer_s4_dd)
        if not parent:
            self.status_label.Text = "Please select a parent layer on the S4 tab."
            self.status_label.TextColor = drawing.Colors.Red
            return

        key_seq = [cb.Text.strip() for cb in self._s4_key_rows]
        key_seq = [k for k in key_seq if k]   # drop blank entries
        if not key_seq:
            self.status_label.Text = "S4 requires at least one attribute key."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        data = calc_s4(parent, key_seq)

        if not data["tree"]:
            self.results_s4.Text = f"  No objects found in any sublayer of '{parent}'.\n"
            self.status_label.Text = "No objects found in sublayers."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        self.results_s4.Text = format_s4(data, parent, key_seq, unit)
        self._last_s4 = data
        self._export_data = {
            "scenario": 4, "unit": unit,
            "params": {"parent": parent, "key_sequence": key_seq},
            "tree": data["tree"],
            "overall_total": data["overall_total"],
        }
        n_warn = len(data["warnings"])
        self.status_label.Text = (
            f"S4  —  '{short_name(parent)}'  |  "
            f"{len(key_seq)} level(s)  |  "
            f"Overall total: {data['overall_total']:,.4f} {unit}"
            + (f"  |  {n_warn} warning(s)" if n_warn else "")
        )
        self.status_label.TextColor = (
            drawing.Colors.Orange if n_warn else drawing.Colors.Green
        )

    # ------------------------------------------------------------------
    # Per-scenario runners — R1 / R2
    # ------------------------------------------------------------------

    def _run_r1(self, unit):
        source_idx = self.r1r2_source_dd.SelectedIndex  # 0=S3 keys, 1=S4 hierarchy

        if source_idx == 0:
            parent  = self._selected_layer(self.parent_layer_dd)
            obj_key = self.obj_key_s3.Text.strip()
            if not parent:
                self.status_label.Text = "R1 uses S3 Parent Layer — select one on the S3 tab."
                self.status_label.TextColor = drawing.Colors.Red
                return
            if not obj_key:
                self.status_label.Text = "R1 uses S3 Object Key — set one on the S3 tab."
                self.status_label.TextColor = drawing.Colors.Orange
                return
        else:
            parent  = self._selected_layer(self.parent_layer_s4_dd)
            key_seq = [cb.Text.strip() for cb in self._s4_key_rows]
            idx     = int(self.r1_level_stepper.Value) - 1   # 0-based
            if not parent:
                self.status_label.Text = "R1 uses S4 Parent Layer — select one on the S4 tab."
                self.status_label.TextColor = drawing.Colors.Red
                return
            if idx < 0 or idx >= len(key_seq) or not key_seq[idx]:
                self.status_label.Text = (
                    f"R1: Level {idx + 1} is undefined in the S4 key sequence."
                )
                self.status_label.TextColor = drawing.Colors.Orange
                return
            obj_key = key_seq[idx]

        room_target_key = self.room_target_key_dd.Text.strip()
        tol  = self.tolerance_stepper.Value / 100.0
        data = calc_r1(parent, obj_key, room_target_key)

        if not data["entries"]:
            self.status_label.Text = "No objects found — check parent layer setting."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        self._r1_entries = data["entries"]
        self._r1_tol     = tol
        self._r1_unit    = unit
        n = len(data["entries"])
        self.chart_r1.Size = drawing.Size(
            max(400, self.chart_r1.Width),
            max(10, n * _CHART_ROW_H + 20)
        )
        self.chart_r1.Invalidate()

        if data["warnings"]:
            self.warn_r1.Text    = "\n".join(data["warnings"])
            self.warn_r1.Visible = True
        else:
            self.warn_r1.Visible = False

        n_warn = len(data["warnings"])
        src_tag = "S3" if source_idx == 0 else "S4"
        self.status_label.Text = (
            f"R1  [{src_tag}]  —  {n} room type(s)  |  Tolerance: {tol*100:.1f}%"
            + (f"  |  {n_warn} warning(s)" if n_warn else "")
        )
        self.status_label.TextColor = (
            drawing.Colors.Orange if n_warn else drawing.Colors.Green
        )

    def _run_r2(self, unit):
        source_idx = self.r1r2_source_dd.SelectedIndex  # 0=S3 keys, 1=S4 hierarchy

        if source_idx == 0:
            parent  = self._selected_layer(self.parent_layer_dd)
            grp_key = self.grp_key_s3.Text.strip()
            if not parent:
                self.status_label.Text = "R2 uses S3 Parent Layer — select one on the S3 tab."
                self.status_label.TextColor = drawing.Colors.Red
                return
            if not grp_key:
                self.status_label.Text = "R2 uses S3 Group Key — set one on the S3 tab."
                self.status_label.TextColor = drawing.Colors.Orange
                return
        else:
            parent  = self._selected_layer(self.parent_layer_s4_dd)
            key_seq = [cb.Text.strip() for cb in self._s4_key_rows]
            idx     = int(self.r2_level_stepper.Value) - 1   # 0-based
            if not parent:
                self.status_label.Text = "R2 uses S4 Parent Layer — select one on the S4 tab."
                self.status_label.TextColor = drawing.Colors.Red
                return
            if idx < 0 or idx >= len(key_seq) or not key_seq[idx]:
                self.status_label.Text = (
                    f"R2: Level {idx + 1} is undefined in the S4 key sequence."
                )
                self.status_label.TextColor = drawing.Colors.Orange
                return
            grp_key = key_seq[idx]

        grp_target_key = self.grp_target_key_dd.Text.strip()
        tol  = self.tolerance_stepper.Value / 100.0
        data = calc_r2(parent, grp_key, grp_target_key)

        if not data["entries"]:
            self.status_label.Text = "No objects found — check parent layer setting."
            self.status_label.TextColor = drawing.Colors.Orange
            return

        self._r2_entries = data["entries"]
        self._r2_tol     = tol
        self._r2_unit    = unit
        n = len(data["entries"])
        self.chart_r2.Size = drawing.Size(
            max(400, self.chart_r2.Width),
            max(10, n * _CHART_ROW_H + 20)
        )
        self.chart_r2.Invalidate()

        if data["warnings"]:
            self.warn_r2.Text    = "\n".join(data["warnings"])
            self.warn_r2.Visible = True
        else:
            self.warn_r2.Visible = False

        n_warn = len(data["warnings"])
        src_tag = "S3" if source_idx == 0 else "S4"
        self.status_label.Text = (
            f"R2  [{src_tag}]  —  {n} group(s)  |  Tolerance: {tol*100:.1f}%"
            + (f"  |  {n_warn} warning(s)" if n_warn else "")
        )
        self.status_label.TextColor = (
            drawing.Colors.Orange if n_warn else drawing.Colors.Green
        )

    # ------------------------------------------------------------------
    # Bullet chart paint handlers
    # ------------------------------------------------------------------

    def _paint_r1(self, sender, e):
        g = e.Graphics
        w = sender.Width
        for i, entry in enumerate(self._r1_entries):
            _draw_bullet_row(g, i, entry, self._r1_tol, self._r1_unit, w)

    def _paint_r2(self, sender, e):
        g = e.Graphics
        w = sender.Width
        for i, entry in enumerate(self._r2_entries):
            _draw_bullet_row(g, i, entry, self._r2_tol, self._r2_unit, w)


# ══════════════════════════════════════════════════════════════════════════════
# Excel export
# ══════════════════════════════════════════════════════════════════════════════

# Styles -------------------------------------------------------------------

_HDR_FONT  = Font(bold=True, color="FFFFFF")
_HDR_FILL  = PatternFill(fill_type="solid", fgColor="2F5496")   # dark blue
_HDR_ALIGN = Alignment(horizontal="center", vertical="center")

_SEC_FONT  = Font(bold=True, color="1F3864")
_SEC_FILL  = PatternFill(fill_type="solid", fgColor="D9E1F2")   # light blue

_TOT_FONT  = Font(bold=True)
_TOT_FILL  = PatternFill(fill_type="solid", fgColor="F2F2F2")   # light grey

_WARN_FONT = Font(bold=True, color="7F6000")
_WARN_FILL = PatternFill(fill_type="solid", fgColor="FFE699")   # amber

_AREA_FMT  = "#,##0.0000"


def _hdr(ws, row, cols):
    """Write a styled header row. Returns the row index + 1."""
    for ci, text in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=text)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.alignment = _HDR_ALIGN
    return row + 1


def _sec(ws, row, text, n_cols=2):
    """Write a section-label row."""
    c = ws.cell(row=row, column=1, value=text)
    c.font = _SEC_FONT
    c.fill = _SEC_FILL
    for ci in range(2, n_cols + 1):
        ws.cell(row=row, column=ci).fill = _SEC_FILL
    return row + 1


def _tot(ws, row, label, value):
    """Write a bold total row with area formatting."""
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = _TOT_FONT
    lc.fill = _TOT_FILL
    vc = ws.cell(row=row, column=2, value=value)
    vc.font          = _TOT_FONT
    vc.fill          = _TOT_FILL
    vc.number_format = _AREA_FMT
    return row + 1


def _warn(ws, row, label, value=None, n_cols=2):
    """Write an amber-highlighted warning row."""
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = _WARN_FONT
    lc.fill = _WARN_FILL
    for ci in range(2, n_cols + 1):
        ws.cell(row=row, column=ci).fill = _WARN_FILL
    if value is not None:
        vc = ws.cell(row=row, column=2, value=value)
        vc.font          = _WARN_FONT
        vc.fill          = _WARN_FILL
        vc.number_format = _AREA_FMT
    return row + 1


def _col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# Per-scenario writers -----------------------------------------------------

def _xl_s1(wb, data, unit):
    name_col = data["params"]["name_key"] or "Object Name"

    ws = wb.create_sheet("Objects")
    _col_widths(ws, [38, 32, 20])
    row = _hdr(ws, 1, ["GUID", name_col, f"Footprint Area ({unit})"])
    for obj in data["objects"]:
        ws.cell(row, 1, obj["guid"])
        ws.cell(row, 2, obj["name"])
        ws.cell(row, 3, obj["area"]).number_format = _AREA_FMT
        row += 1

    ws2 = wb.create_sheet("Summary")
    _col_widths(ws2, [36, 22])
    row = _hdr(ws2, 1, ["S1 — Selected Objects", f"[{unit}]"])
    pairs = [
        ("Name Key used",        data["params"]["name_key"] or "(object name / GUID)"),
        ("Object count",         len(data["objects"])),
        ("Total footprint area", sum(o["area"] for o in data["objects"])),
    ]
    for label, value in pairs:
        lc = ws2.cell(row, 1, label)
        lc.font = Font(bold=True)
        vc = ws2.cell(row, 2, value)
        if isinstance(value, float):
            vc.number_format = _AREA_FMT
        row += 1


def _xl_s2(wb, data, unit):
    obj_col = data["params"]["obj_key"] or "Object Name"
    layer   = data["params"]["layer_name"]
    raw_sum = sum(o["area"] for o in data["objects"])

    ws = wb.create_sheet("Objects")
    _col_widths(ws, [38, 30, 32, 20])
    row = _hdr(ws, 1, ["GUID", "Layer", obj_col, f"Footprint Area ({unit})"])
    for obj in data["objects"]:
        ws.cell(row, 1, obj["guid"])
        ws.cell(row, 2, layer)
        ws.cell(row, 3, obj["name"])
        ws.cell(row, 4, obj["area"]).number_format = _AREA_FMT
        row += 1

    ws2 = wb.create_sheet("Summary")
    _col_widths(ws2, [38, 22])
    row = _hdr(ws2, 1, ["S2 — By Layer", f"[{unit}]"])
    kv_rows = [
        ("Layer",                              layer),
        ("Object Key used",                    data["params"]["obj_key"] or "(object name / GUID)"),
        ("Object count",                       len(data["objects"])),
        (f"Sum of individual areas ({unit})",  raw_sum),
        (f"Combined footprint ({unit})",       data["total"]),
        ("Boolean Union succeeded",            "Yes" if data["union_ok"] else "No — sum shown"),
    ]
    for label, value in kv_rows:
        lc = ws2.cell(row, 1, label)
        lc.font = Font(bold=True)
        vc = ws2.cell(row, 2, value)
        if isinstance(value, float):
            vc.number_format = _AREA_FMT
        row += 1

    overlap = raw_sum - data["total"]
    if overlap > 1e-6:
        row += 1
        row = _sec(ws2, row, "[!] Overlap Warning", n_cols=2)
        row = _warn(ws2, row, "Overlapping area (sum - total)", overlap)
        row = _warn(ws2, row, "Some objects share footprint area. Verify whether double-counting is intentional.")


def _xl_s3(wb, data, unit):
    params  = data["params"]
    obj_col = params["obj_key"] or "Object Name"
    grp_col = params["grp_key"] or "Group"
    parent  = params["parent"]
    has_grp = bool(params["grp_key"])

    ws = wb.create_sheet("Objects")
    _col_widths(ws, [38, 28, 28, 30, 28, 20])
    row = _hdr(ws, 1, ["GUID", "Parent Layer", "Level", obj_col, grp_col, f"Footprint Area ({unit})"])
    for sl, sl_data in data["sublayers"].items():
        level = short_name(sl)
        for obj in sl_data["objects"]:
            ws.cell(row, 1, obj["guid"])
            ws.cell(row, 2, parent)
            ws.cell(row, 3, level)
            ws.cell(row, 4, obj["name"])
            ws.cell(row, 5, obj["group"])
            ws.cell(row, 6, obj["area"]).number_format = _AREA_FMT
            row += 1

    ws2 = wb.create_sheet("Summary")
    _col_widths(ws2, [30, 28, 22])
    row = _hdr(ws2, 1, ["S3 — Layer Hierarchy", "", f"[{unit}]"])

    for label, value in [
        ("Parent Layer", parent),
        ("Object Key",   params["obj_key"] or "(object name / GUID)"),
        ("Group Key",    params["grp_key"] or "—"),
    ]:
        ws2.cell(row, 1, label).font = Font(bold=True)
        ws2.cell(row, 2, value)
        row += 1
    row += 1

    row = _sec(ws2, row, "Combined Footprint by Level", n_cols=2)
    ws2.cell(row, 1, "Level").font           = Font(bold=True)
    ws2.cell(row, 2, f"Area ({unit})").font  = Font(bold=True)
    row += 1
    for sl, sl_data in data["sublayers"].items():
        note = "" if sl_data["union_ok"] else " [union failed]"
        ws2.cell(row, 1, short_name(sl) + note)
        ws2.cell(row, 2, sl_data["total"]).number_format = _AREA_FMT
        row += 1
    row = _tot(ws2, row, "Grand Total", data["overall_total"])
    row += 1

    if has_grp:
        row = _sec(ws2, row, f"Combined Footprint by Level and {grp_col}", n_cols=3)
        ws2.cell(row, 1, "Level").font          = Font(bold=True)
        ws2.cell(row, 2, grp_col).font          = Font(bold=True)
        ws2.cell(row, 3, f"Area ({unit})").font = Font(bold=True)
        row += 1
        for sl, sl_data in data["sublayers"].items():
            level = short_name(sl)
            for gv, ga in sorted(sl_data["group_totals"].items()):
                ws2.cell(row, 1, level)
                ws2.cell(row, 2, gv)
                ws2.cell(row, 3, ga).number_format = _AREA_FMT
                row += 1

    levels_with_overlap = []
    for sl, sl_data in data["sublayers"].items():
        individual_sum = sum(o["area"] for o in sl_data["objects"])
        total_overlap  = individual_sum - sl_data["total"]
        if total_overlap > 1e-6:
            group_sum   = sum(sl_data["group_totals"].values()) if sl_data["group_totals"] else individual_sum
            cross_group = group_sum - sl_data["total"]
            levels_with_overlap.append((short_name(sl), total_overlap, cross_group))

    if levels_with_overlap:
        row += 1
        n_warn_cols = 3 if has_grp else 2
        row = _sec(ws2, row, "[!] Overlap Warnings by Level", n_cols=n_warn_cols)
        ws2.cell(row, 1, "Level").font                         = Font(bold=True)
        ws2.cell(row, 2, f"Overlapping Area ({unit})").font   = Font(bold=True)
        if has_grp:
            ws2.cell(row, 3, f"of which Cross-group ({unit})").font = Font(bold=True)
        row += 1
        for level_name, total_ov, cross_ov in levels_with_overlap:
            row = _warn(ws2, row, level_name, total_ov, n_cols=n_warn_cols)
            if has_grp and cross_ov > 1e-6:
                vc = ws2.cell(row - 1, 3, cross_ov)
                vc.font          = _WARN_FONT
                vc.fill          = _WARN_FILL
                vc.number_format = _AREA_FMT
        row += 1
        row = _warn(ws2, row,
                    "Some objects share footprint area. Verify whether double-counting is intentional.",
                    n_cols=n_warn_cols)


def _xl_s4(wb, data, unit):
    """
    S4 Custom Aggregation Excel export.
    Sheet 'Leaf Data': flat table — one row per unique key path (leaf), with
    one column per key level plus an area column. Useful for pivot tables.
    Sheet 'Tree Summary': indented hierarchy showing subtotals at each node.
    """
    params   = data["params"]
    parent   = params["parent"]
    key_seq  = params["key_sequence"]
    depth    = len(key_seq)
    key_path = " > ".join(key_seq)

    # ── Sheet 1: Leaf Data ───────────────────────────────────────────
    ws = wb.create_sheet("Leaf Data")
    col_widths = [22] * depth + [20]
    _col_widths(ws, col_widths)
    header_cols = list(key_seq) + [f"Area ({unit})"]
    row = _hdr(ws, 1, header_cols)

    def write_leaves(node, path_so_far):
        nonlocal row
        for val in sorted(node):
            entry = node[val]
            current_path = path_so_far + [val]
            if entry["children"]:
                write_leaves(entry["children"], current_path)
            else:
                # Leaf row: fill each key column
                for ci, pv in enumerate(current_path, 1):
                    ws.cell(row, ci, pv)
                ws.cell(row, depth + 1, entry["area"]).number_format = _AREA_FMT
                row += 1

    write_leaves(data["tree"], [])

    # ── Sheet 2: Tree Summary ────────────────────────────────────────
    ws2 = wb.create_sheet("Tree Summary")
    _col_widths(ws2, [52, 22])
    row2 = _hdr(ws2, 1, ["S4 — Custom Aggregation", f"Area ({unit})"])

    # Header info rows
    for lbl, val in [("Parent Layer", parent), ("Key Hierarchy", key_path)]:
        ws2.cell(row2, 1, lbl).font = Font(bold=True)
        ws2.cell(row2, 2, val)
        row2 += 1
    row2 += 1

    def write_tree_rows(node, level):
        nonlocal row2
        for val in sorted(node):
            entry = node[val]
            indent = "    " * level
            is_leaf = not entry["children"]
            label = f"{indent}{'▸ ' if not is_leaf else '  '}{val}"
            lc = ws2.cell(row2, 1, label)
            vc = ws2.cell(row2, 2, entry["area"])
            vc.number_format = _AREA_FMT
            if not is_leaf:
                lc.font = Font(bold=True)
                lc.fill = _SEC_FILL
                vc.font = Font(bold=True)
                vc.fill = _SEC_FILL
            row2 += 1
            if entry["children"]:
                write_tree_rows(entry["children"], level + 1)

    write_tree_rows(data["tree"], 0)
    row2 += 1

    # Overall total
    lc = ws2.cell(row2, 1, "OVERALL TOTAL")
    lc.font = _TOT_FONT
    lc.fill = _TOT_FILL
    vc = ws2.cell(row2, 2, data["overall_total"])
    vc.font          = _TOT_FONT
    vc.fill          = _TOT_FILL
    vc.number_format = _AREA_FMT


def export_to_excel(data, filepath):
    """Write calculation results to an Excel workbook with Objects + Summary sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    s    = data["scenario"]
    unit = data["unit"]

    if s == 1:
        _xl_s1(wb, data, unit)
    elif s == 2:
        _xl_s2(wb, data, unit)
    elif s == 3:
        _xl_s3(wb, data, unit)
    else:
        _xl_s4(wb, data, unit)

    wb.save(filepath)


# ══════════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════════

def main():
    form = LinderoForm()
    form.Show()


if __name__ == "__main__":
    main()
