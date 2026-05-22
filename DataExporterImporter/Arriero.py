#! python3
# r: openpyxl
# -*- coding: utf-8 -*-
# __title__ = "Arriero"
# __doc__ = """Version = 0.5
# Date    = 2026-02-14
# Author: Aquelon - aquelon@pm.me
# _____________________________________________________________________
# Description:
# DataExporterImporter for Rhino 8
# Exports and imports object User Keys/Values to/from Excel files using GUID tracking.
# _____________________________________________________________________
# How-to:
# -> Run the script in Rhino 8 (RunPythonScript)
# -> A window opens with Export and Import buttons and import options
# -> To export: Click "Export Data", select objects, choose save location
# -> To import: Configure options (backup, create keys, empty cells), click "Import Data"
# -> Select the Excel file, review the summary report
# -> Tip: Use "Apply only to pre-selected objects" to limit new key creation scope
# _____________________________________________________________________
# Last update:
# - [14.02.2026] - 0.5 RELEASE
# _____________________________________________________________________
# To-Do:
# - UI needs improvement, it is functional at the moment, but default look.
# - Default folder locations need to be updated to neutral locations (desktop or documents) instead of script folder.
# - Add function that erase keys on objects if they are not present in Excel file, this will allow to use the tool for cleaning up objects by erasing the columns in Excel.
# Update the logic of the import with empty cells, it seems tha is bit mixed right now. If the cells are emptied in an object, it is possible to erase the key in that object, but this logic may not be transparent for the user.
# _____________________________________________________________________

import rhinoscriptsyntax as rs
import Rhino
import System
import datetime
import os

# ============================================================================
# PERSISTENT PREFERENCES (last-used folder per action)
# ============================================================================

import sys as _sys, os as _os
_rg_root = _os.path.normpath(_os.path.join(_os.path.dirname(_os.path.abspath(__file__)), ".."))
if _rg_root not in _sys.path:
    _sys.path.insert(0, _rg_root)
from ui import theme as _t
import importlib as _importlib
_importlib.reload(_t)

# Import Eto forms and drawing
import Eto.Drawing as drawing
import Eto.Forms as forms

# Try to import openpyxl for Excel operations
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available. Please install it using: pip install openpyxl")



class DataExporterImporterGUI(forms.Dialog[bool]):
    """Main GUI for the Data Exporter/Importer tool using Eto.Forms"""
    
    def __init__(self):
        super(DataExporterImporterGUI, self).__init__()
        self.Title = "Rhino Data Exporter/Importer"
        self.BackgroundColor = _t.BG
        self.Resizable = False
        self.ClientSize = drawing.Size(480, 520)
        self.operation = None
        
        self._create_controls()
        self._create_layout()
        
    def _create_controls(self):
        # Section Label
        self.op_header = _t.section_header("Select Operation:")
        
        # Export Button
        self.export_btn = forms.Button()
        self.export_btn.Text = "1. Export Data (Select Objects → Excel)"
        self.export_btn.Font = _t.F_SANS_B
        self.export_btn.BackgroundColor = _t.BTN_CALC
        self.export_btn.Height = 40
        self.export_btn.Click += self.on_export_click
        
        # Import Button
        self.import_btn = forms.Button()
        self.import_btn.Text = "2. Import Data (Excel → Update Objects)"
        self.import_btn.Font = _t.F_SANS_B
        self.import_btn.BackgroundColor = _t.BTN_CALC
        self.import_btn.Height = 40
        self.import_btn.Click += self.on_import_click
        
        # Options Header
        self.options_header = _t.section_header("Import Options:")
        
        # Checkboxes
        self.backup_check = forms.CheckBox()
        self.backup_check.Text = "Create backup before import"
        self.backup_check.Checked = True
        
        self.create_keys_check = forms.CheckBox()
        self.create_keys_check.Text = "Create missing keys from Excel columns"
        self.create_keys_check.Checked = True
        self.create_keys_check.CheckedChanged += self.on_create_keys_changed
        
        # Radio buttons (grouped)
        self.new_keys_all_radio = forms.RadioButton()
        self.new_keys_all_radio.Text = "Apply to all objects in Excel file"
        self.new_keys_all_radio.Checked = True
        
        self.new_keys_selected_radio = forms.RadioButton(self.new_keys_all_radio)
        self.new_keys_selected_radio.Text = "Apply only to pre-selected objects"
        
        self.update_empty_check = forms.CheckBox()
        self.update_empty_check.Text = "Update values when Excel cell is empty"
        self.update_empty_check.Checked = False
        self.update_empty_check.CheckedChanged += self.on_update_empty_changed
        
        # Placeholder text & label
        self.placeholder_label = forms.Label()
        self.placeholder_label.Text = "Placeholder value (leave blank to DELETE the key):"
        self.placeholder_label.Font = _t.F_SANS_S
        self.placeholder_label.TextColor = _t.TEXT_MUTED
        self.placeholder_label.VerticalAlignment = forms.VerticalAlignment.Center
        self.placeholder_label.Enabled = False
        
        self.placeholder_text = forms.TextBox()
        self.placeholder_text.Text = "-"
        self.placeholder_text.Width = 60
        self.placeholder_text.Enabled = False
        
        # Close button
        self.close_btn = forms.Button()
        self.close_btn.Text = "Cancel"
        self.close_btn.Font = _t.F_SANS_B
        self.close_btn.BackgroundColor = _t.BTN_CLEAR
        self.close_btn.Click += self.on_close_click
        
    def _create_layout(self):
        # Create a beautiful options container
        options_layout = forms.DynamicLayout()
        options_layout.Spacing = drawing.Size(5, 8)
        options_layout.AddRow(self.backup_check)
        options_layout.AddRow(self.create_keys_check)
        
        # Indented radio options
        radio_layout = forms.DynamicLayout()
        radio_layout.Spacing = drawing.Size(5, 5)
        radio_container = forms.Panel()
        radio_container.Padding = drawing.Padding(20, 2, 0, 2)
        radio_layout.AddRow(self.new_keys_all_radio)
        radio_layout.AddRow(self.new_keys_selected_radio)
        radio_container.Content = radio_layout
        options_layout.AddRow(radio_container)
        
        options_layout.AddRow(self.update_empty_check)
        
        # Placeholder horizontal layout
        ph_layout = forms.DynamicLayout()
        ph_layout.Spacing = drawing.Size(10, 5)
        ph_layout.AddRow(self.placeholder_label, self.placeholder_text)
        
        ph_container = forms.Panel()
        ph_container.Padding = drawing.Padding(20, 2, 0, 2)
        ph_container.Content = ph_layout
        options_layout.AddRow(ph_container)
        
        # Options Panel
        options_panel = forms.Panel()
        options_panel.BackgroundColor = _t.PANEL
        options_panel.Padding = drawing.Padding(15)
        options_panel.Content = options_layout
        
        # Main Layout
        main_layout = forms.DynamicLayout()
        main_layout.Spacing = drawing.Size(10, 10)
        main_layout.Padding = drawing.Padding(20)
        
        main_layout.AddRow(self.op_header)
        main_layout.AddRow(self.export_btn)
        main_layout.AddRow(self.import_btn)
        main_layout.AddRow(None) # Spacer
        main_layout.AddRow(self.options_header)
        main_layout.AddRow(options_panel)
        main_layout.AddRow(None) # Spacer
        main_layout.AddRow(self.close_btn)
        
        self.Content = main_layout
        
    def on_create_keys_changed(self, sender, event):
        enabled = self.create_keys_check.Checked
        self.new_keys_all_radio.Enabled = enabled
        self.new_keys_selected_radio.Enabled = enabled
        
    def on_update_empty_changed(self, sender, event):
        enabled = self.update_empty_check.Checked
        self.placeholder_label.Enabled = enabled
        self.placeholder_text.Enabled = enabled
        
    def on_export_click(self, sender, event):
        if not EXCEL_AVAILABLE:
            forms.MessageBox.Show(
                "openpyxl library is not available.\nPlease install it using: pip install openpyxl",
                "Missing Dependency",
                forms.MessageBoxButtons.OK,
                forms.MessageBoxType.Error
            )
            return
        self.operation = "export"
        self.Close(True)
        
    def on_import_click(self, sender, event):
        if not EXCEL_AVAILABLE:
            forms.MessageBox.Show(
                "openpyxl library is not available.\nPlease install it using: pip install openpyxl",
                "Missing Dependency",
                forms.MessageBoxButtons.OK,
                forms.MessageBoxType.Error
            )
            return
        self.operation = "import"
        self.Close(True)
        
    def on_close_click(self, sender, event):
        self.Close(False)


def export_data_to_excel():
    """
    Export selected objects' User Keys and Values to Excel file
    """
    # Prompt user to select objects
    objects = rs.GetObjects("Select objects to export data", preselect=True)
    if not objects:
        print("No objects selected. Export cancelled.")
        return
    
    print(f"Selected {len(objects)} objects for export.")
    
    # Collect all unique keys across all objects
    all_keys = set()
    object_data = []
    
    for obj in objects:
        guid = str(obj)
        keys = rs.GetUserText(obj)
        
        obj_dict = {"GUID": guid}
        
        if keys:
            for key in keys:
                value = rs.GetUserText(obj, key)
                obj_dict[key] = value if value else ""
                all_keys.add(key)
        
        object_data.append(obj_dict)
    
    # Sort keys alphabetically for consistent column order
    sorted_keys = sorted(all_keys)
    column_headers = ["GUID"] + sorted_keys
    
    print(f"Found {len(sorted_keys)} unique keys across all objects.")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rhino Object Data"
    
    # Write headers
    for col_idx, header in enumerate(column_headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write data rows
    for row_idx, obj_dict in enumerate(object_data, start=2):
        for col_idx, header in enumerate(column_headers, start=1):
            value = obj_dict.get(header, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Prompt user to select save location
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
    default_filename = f"RhinoData_{timestamp}.xlsx"

    folder = _t.prefs_get('arriero_export')
    filepath = rs.SaveFileName(
        title="Save Excel File",
        filter="Excel Files (*.xlsx)|*.xlsx|All Files (*.*)|*.*||",
        folder=folder,
        filename=default_filename,
        extension="xlsx"
    )
    if not filepath:
        print("Export cancelled by user.")
        return

    _t.prefs_set('arriero_export', filepath)

    # Save workbook
    try:
        wb.save(filepath)
        print(f"\nExport successful!")
        print(f"File saved: {filepath}")
        print(f"Exported {len(object_data)} objects with {len(sorted_keys)} keys.")

        forms.MessageBox.Show(
            f"Export successful!\n\n"
            f"Objects: {len(object_data)}\n"
            f"Keys: {len(sorted_keys)}\n\n"
            f"File: {os.path.basename(filepath)}",
            "Export Complete",
            forms.MessageBoxButtons.OK,
            forms.MessageBoxType.Information
        )
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        forms.MessageBox.Show(
            f"Error saving Excel file:\n{e}",
            "Export Error",
            forms.MessageBoxButtons.OK,
            forms.MessageBoxType.Error
        )


def import_data_from_excel(create_backup, create_missing_keys,
                           update_empty, placeholder, only_selected):
    """
    Import data from Excel file and update objects based on GUID

    Args:
        create_backup (bool): Create backup before import
        create_missing_keys (bool): Create keys that don't exist on objects
        update_empty (bool): Update values when Excel cell is empty
        placeholder (str): Placeholder value for empty cells (empty string means delete key)
        only_selected (bool): Only apply new keys to pre-selected objects
    """
    # Get pre-selected objects if only_selected is True
    selected_guids = None
    if only_selected and create_missing_keys:
        selected_objects = rs.SelectedObjects()
        if not selected_objects:
            forms.MessageBox.Show(
                "No objects are selected.\n\n"
                "Please select objects first if you want to use\n"
                "'Apply only to pre-selected objects' option.",
                "No Selection",
                forms.MessageBoxButtons.OK,
                forms.MessageBoxType.Warning
            )
            return
        selected_guids = set(str(obj) for obj in selected_objects)
        print(f"Pre-selected {len(selected_guids)} objects for new key creation.")

    # Open file dialog to select Excel file
    folder = _t.prefs_get('arriero_import')
    filepath = rs.OpenFileName(
        title="Select Excel file to import",
        filter="Excel Files (*.xlsx)|*.xlsx|All Files (*.*)|*.*||",
        folder=folder
    )
    if not filepath:
        print("Import cancelled.")
        return

    _t.prefs_set('arriero_import', filepath)
    print(f"Selected file: {filepath}")
    
    # Create backup if requested
    if create_backup:
        result = forms.MessageBox.Show(
            "Do you want to create a backup before importing?\n\n"
            "This will export all current object data to a new Excel file.",
            "Create Backup",
            forms.MessageBoxButtons.YesNo,
            forms.MessageBoxType.Question
        )

        if result == forms.MessageBoxResult.Yes:
            print("\nCreating backup before import...")
            # Get all objects in document for backup
            all_objects = rs.AllObjects()
            if all_objects:
                # Select all objects for backup
                rs.SelectObjects(all_objects)
                export_data_to_excel()
                rs.UnselectAllObjects()
                print("Backup created.")
            else:
                print("No objects found for backup.")
    
    # Load Excel file
    try:
        wb = load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        forms.MessageBox.Show(
            f"Error loading Excel file:\n{e}",
            "Import Error",
            forms.MessageBoxButtons.OK,
            forms.MessageBoxType.Error
        )
        return
    
    # Read headers (first row)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value))
    
    if not headers or headers[0] != "GUID":
        print("Error: First column must be 'GUID'")
        forms.MessageBox.Show(
            "Invalid Excel format.\nFirst column must be 'GUID'.",
            "Import Error",
            forms.MessageBoxButtons.OK,
            forms.MessageBoxType.Error
        )
        return
    
    # Keys are all columns except GUID
    keys = headers[1:]
    print(f"\nFound {len(keys)} keys in Excel file.")
    
    # Process each row
    updated_count = 0
    not_found_count = 0
    not_found_guids = []
    keys_created = 0
    keys_updated = 0
    keys_deleted = 0
    skipped_not_selected = 0
    
    for row_idx in range(2, ws.max_row + 1):
        guid_str = ws.cell(row=row_idx, column=1).value
        
        if not guid_str:
            continue
        
        # Try to find object by GUID
        try:
            guid = System.Guid(str(guid_str))
            obj = rs.coerceguid(guid)
            
            if not rs.IsObject(obj):
                not_found_count += 1
                not_found_guids.append(guid_str)
                continue

            # Check if this object can receive new keys
            can_create_keys = create_missing_keys
            object_skipped_selection = False
            if can_create_keys and selected_guids is not None:
                # Only allow key creation if object is in selected set
                if guid_str not in selected_guids:
                    can_create_keys = False
                    object_skipped_selection = True

            # Process each key/value pair
            for col_idx, key in enumerate(keys, start=2):
                excel_value = ws.cell(row=row_idx, column=col_idx).value
                
                # Check if value is None or empty string
                is_empty = excel_value is None or str(excel_value).strip() == ""
                
                if is_empty and not update_empty:
                    # Skip empty cells if update_empty is False
                    continue
                
                # Get current value on object
                current_value = rs.GetUserText(obj, key)
                key_exists = current_value is not None
                
                if is_empty and update_empty:
                    # Empty cell with update_empty enabled
                    if placeholder:
                        # Set to placeholder
                        if not key_exists and not can_create_keys:
                            # Skip if key doesn't exist and we can't create it
                            if object_skipped_selection:
                                skipped_not_selected += 1
                            continue
                        rs.SetUserText(obj, key, placeholder)
                        if key_exists:
                            keys_updated += 1
                        else:
                            keys_created += 1
                    else:
                        # Delete the key
                        if key_exists:
                            rs.SetUserText(obj, key, None)
                            keys_deleted += 1
                else:
                    # Non-empty value
                    value_str = str(excel_value)

                    if not key_exists and not can_create_keys:
                        # Skip if key doesn't exist and we can't create it
                        if object_skipped_selection:
                            skipped_not_selected += 1
                        continue

                    rs.SetUserText(obj, key, value_str)

                    if key_exists:
                        keys_updated += 1
                    else:
                        keys_created += 1
            
            updated_count += 1
            
        except Exception as e:
            print(f"Error processing GUID {guid_str}: {e}")
            not_found_count += 1
            not_found_guids.append(guid_str)
            continue
    
    # Show summary
    summary_msg = (
        f"Import Complete!\n\n"
        f"Objects updated: {updated_count}\n"
        f"Keys created: {keys_created}\n"
        f"Keys updated: {keys_updated}\n"
        f"Keys deleted: {keys_deleted}\n"
    )

    if skipped_not_selected > 0:
        summary_msg += f"Keys skipped (not in selection): {skipped_not_selected}\n"

    if not_found_count > 0:
        summary_msg += f"\nGUIDs not found: {not_found_count}\n"
        if len(not_found_guids) <= 5:
            summary_msg += "\n".join([f"  - {g}" for g in not_found_guids])
        else:
            summary_msg += "\n".join([f"  - {g}" for g in not_found_guids[:5]])
            summary_msg += f"\n  ... and {len(not_found_guids) - 5} more"
    
    print("\n" + "="*50)
    print(summary_msg)
    print("="*50)
    
    forms.MessageBox.Show(
        summary_msg,
        "Import Summary",
        forms.MessageBoxButtons.OK,
        forms.MessageBoxType.Information
    )


def main():
    """Main function to run the Data Exporter/Importer tool"""
    
    if not EXCEL_AVAILABLE:
        print("="*60)
        print("ERROR: openpyxl library is not installed")
        print("="*60)
        print("\nThis script requires the openpyxl library to work with Excel files.")
        print("Please install it using the following command:\n")
        print("  pip install openpyxl\n")
        print("If you're using Rhino 8's Python 3, you may need to:")
        print("  1. Open Windows Command Prompt (cmd)")
        print("  2. Navigate to Rhino's Python directory")
        print("  3. Run: python -m pip install openpyxl")
        print("="*60)
        return
    
    # Show GUI
    gui = DataExporterImporterGUI()
    result = gui.ShowModal(Rhino.UI.RhinoEtoApp.MainWindow)
    
    if result:
        operation = gui.operation

        if operation == "export":
            export_data_to_excel()

        elif operation == "import":
            create_backup = gui.backup_check.Checked
            create_missing_keys = gui.create_keys_check.Checked
            update_empty = gui.update_empty_check.Checked
            placeholder = gui.placeholder_text.Text if update_empty else ""
            only_selected = gui.new_keys_selected_radio.Checked

            import_data_from_excel(
                create_backup,
                create_missing_keys,
                update_empty,
                placeholder,
                only_selected
            )


if __name__ == "__main__":
    main()
